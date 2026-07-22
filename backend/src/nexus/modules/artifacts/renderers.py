from __future__ import annotations

import csv
import html
import io
import json
import re
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from nexus.shared.domain.errors import ValidationError

EVIDENCE_MARKER_RE = re.compile(r"\[evidence:([0-9a-f-]{36})\]", re.IGNORECASE)
MARKDOWN_HEADING_RE = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
ORDERED_LIST_RE = re.compile(r"^\s*(\d+)\.\s+(.+?)\s*$")
BULLET_LIST_RE = re.compile(r"^\s*[-*]\s+(.+?)\s*$")
DEFAULT_ID_PREFIX_LENGTH = 8
MIN_COLLISION_TAIL_LENGTH = 4
MAX_COLLISION_TAIL_LENGTH = 12


@dataclass(frozen=True)
class ArtifactRenderMetadata:
    artifact_id: str | None = None
    artifact_type: str | None = None
    bound_evidence_count: int | None = None
    content_block_count: int | None = None
    coverage_percent: int | None = None
    pending_refresh_count: int | None = None
    revision_id: str | None = None
    revision_no: int | None = None
    status: str | None = None
    supported_block_count: int | None = None
    updated_at: object | None = None


@dataclass(frozen=True)
class _CitationReference:
    evidence_id: str
    label: str
    short_id: str


@dataclass(frozen=True)
class _SourceReceipt:
    detail: str
    evidence_id: str
    label: str
    locator: str
    short_id: str
    source: str


@dataclass(frozen=True)
class _DocumentOutlineEntry:
    anchor_id: str
    kind: str
    level: int
    title: str


@dataclass(frozen=True)
class _CitationContext:
    references: dict[str, _CitationReference]

    def reference(self, evidence_id: object) -> _CitationReference | None:
        key = _normalize_evidence_id(evidence_id)
        return self.references.get(key) if key else None


def render_artifact(
    document: dict[str, Any],
    format_name: str,
    metadata: ArtifactRenderMetadata | None = None,
) -> tuple[bytes, str, str]:
    """Render one canonical revision without creating a competing source of truth."""
    if format_name == "json":
        return (
            json.dumps(document, ensure_ascii=False, indent=2).encode("utf-8"),
            "application/json",
            "json",
        )
    metadata = metadata or ArtifactRenderMetadata()
    markdown = artifact_markdown(document, metadata)
    if format_name == "markdown":
        return markdown.encode("utf-8"), "text/markdown; charset=utf-8", "md"
    if format_name == "html":
        body = _artifact_html(document, metadata)
        return body.encode("utf-8"), "text/html; charset=utf-8", "html"
    if format_name == "pdf":
        return _artifact_pdf(document, metadata), "application/pdf", "pdf"
    if format_name == "csv":
        return _artifact_csv(document), "text/csv; charset=utf-8", "csv"
    if format_name == "xlsx":
        return (
            _artifact_xlsx(document, metadata),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    raise ValidationError("Unsupported Artifact render format", details={"format": format_name})


def artifact_markdown(
    document: dict[str, Any],
    metadata: ArtifactRenderMetadata | None = None,
) -> str:
    citations = _citation_context(document)
    lines: list[str] = []
    metadata_lines = _markdown_delivery_metadata(
        document,
        metadata or ArtifactRenderMetadata(),
        citations,
    )
    if metadata_lines:
        lines.extend(metadata_lines)
        lines.append("")
    for block in _blocks(document):
        block_type = block.get("type")
        if block_type == "heading":
            lines.append(
                f"{'#' * max(1, min(6, int(block.get('level', 1))))} {block.get('text', '')}"
            )
        elif block_type == "paragraph":
            lines.append(_markdown_text_with_citations(str(block.get("text", "")), citations))
        elif block_type == "evidence_list":
            receipts = _source_receipts(block, citations)
            if receipts:
                lines.append("## Source receipts")
                for receipt in receipts:
                    parts = [
                        f"[{receipt.label}]",
                        receipt.source,
                        receipt.locator,
                        receipt.detail,
                        f"Evidence {receipt.short_id}",
                    ]
                    lines.append("- " + " · ".join(part for part in parts if part))
        elif block_type == "table":
            columns, rows = _table_data(block)
            lines.append("| " + " | ".join(map(str, columns)) + " |")
            lines.append("| " + " | ".join("---" for _ in columns) + " |")
            for row in rows:
                lines.append("| " + " | ".join(str(value) for value in row) + " |")
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def _artifact_html(document: dict[str, Any], metadata: ArtifactRenderMetadata) -> str:
    citations = _citation_context(document)
    outline = _document_outline(document, citations)
    heading_index = 0
    parts = [
        '<!doctype html><html><head><meta charset="utf-8"><title>',
        html.escape(str(document.get("title", "Artifact"))),
        "</title>",
        "<style>",
        "body{margin:0;padding:48px;background:#f4f1e8;color:#1e2a27;"
        "font:16px/1.65 Georgia,'Times New Roman',serif;}"
        "main{max-width:860px;margin:0 auto;padding:56px;background:#fffdf7;"
        "border:1px solid #d8d0bf;box-shadow:0 24px 70px rgba(31,42,39,.12);}"
        "h1,h2,h3,h4,h5,h6{font-family:Georgia,'Times New Roman',serif;letter-spacing:-.025em;}"
        ".delivery-cover{margin:0 0 42px;padding:22px;background:#f5f8f5;border:1px solid #cbd8d2;"
        "border-radius:6px 22px 22px 22px;}.eyebrow{margin:0 0 8px;color:#60716d;"
        "font:800 10px/1.2 ui-monospace,monospace;letter-spacing:.12em;text-transform:uppercase;}"
        ".delivery-title{margin:0;color:#17221f;font-size:30px;line-height:1.12;"
        "letter-spacing:-.04em;}"
        ".delivery-meta{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));"
        "gap:8px;margin:18px 0 0;}.delivery-meta div{padding:10px;background:#fffdf7;"
        "border:1px solid #d3ded9;border-radius:10px;}.delivery-meta dt{color:#60716d;"
        "font:800 9px/1 ui-monospace,monospace;text-transform:uppercase;}"
        ".delivery-meta dd{margin:5px 0 0;}"
        ".delivery-map{margin:-24px 0 38px;padding:16px 18px;background:#fffdf7;"
        "border:1px dashed #c5d2cc;border-radius:14px;}.delivery-map h2{margin:0 0 10px;"
        "font-size:16px;}.delivery-map ol{display:grid;"
        "grid-template-columns:repeat(auto-fit,minmax(180px,1fr));"
        "gap:7px 16px;margin:0;padding-left:18px;}.delivery-map a{color:#1c5d4f;"
        "font:800 11px/1.35 ui-monospace,monospace;"
        "text-decoration:none;}.delivery-map a:hover{text-decoration:underline;}"
        "p{margin:0 0 14px;}hr{height:1px;margin:28px 0;border:0;background:#d8d0bf;}"
        "ul,ol{margin:0 0 18px 24px;padding:0;}li{margin:6px 0;}code{padding:1px 4px;"
        "background:#edf3ef;border:1px solid #d8d0bf;border-radius:5px;"
        "font:12px/1.4 ui-monospace,monospace;}"
        "table{width:100%;border-collapse:collapse;margin:22px 0;font-size:13px;}"
        "th,td{padding:9px 11px;border:1px solid #d8d0bf;text-align:left;vertical-align:top;}"
        "th{background:#edf3ef;font-size:11px;text-transform:uppercase;letter-spacing:.06em;}"
        ".citation{display:inline-grid;place-items:center;min-width:22px;height:18px;margin:0 2px;"
        "padding:0 5px;border:1px solid #b9d8cd;border-radius:999px;background:#edf6f2;"
        "color:#173f37;font:800 9px/1 ui-monospace,monospace;text-decoration:none;"
        "vertical-align:super;}"
        ".sources{margin-top:42px;padding:18px;background:#f5f8f5;border:1px solid #cbd8d2;"
        "border-radius:5px 16px 16px 16px;}"
        ".sources h2{margin:0 0 12px;font-size:20px;}"
        ".sources ol{display:grid;gap:8px;margin:0;padding-left:24px;}"
        ".sources li{padding:10px;background:#fffdf7;border:1px solid #d3ded9;border-radius:10px;}"
        ".sources small{display:block;color:#64736f;font:10px/1.45 ui-monospace,monospace;}"
        "@media print{body{padding:0;background:white;}main{max-width:none;margin:0;padding:0;"
        "border:0;box-shadow:none;}.delivery-cover,.delivery-map,.sources{break-inside:avoid;"
        "box-shadow:none;}a{color:inherit;text-decoration:none;}.citation{border-color:#9fb8af;"
        "background:white;}table{break-inside:auto;}tr{break-inside:avoid;}}"
        "</style></head><body><main>",
    ]
    parts.append(_html_delivery_cover(document, metadata, citations))
    parts.append(_html_document_map(outline))
    for block in _blocks(document):
        block_type = block.get("type")
        if block_type == "heading":
            level = max(1, min(6, int(block.get("level", 1))))
            anchor = _outline_anchor_at(outline, heading_index)
            heading_index += 1
            anchor_attr = f' id="{html.escape(anchor)}"' if anchor else ""
            parts.append(
                f"<h{level}{anchor_attr}>{html.escape(str(block.get('text', '')))}</h{level}>"
            )
        elif block_type == "paragraph":
            rendered, heading_index = _html_markdown_blocks(
                str(block.get("text", "")),
                citations,
                outline,
                heading_index,
            )
            parts.append(rendered)
        elif block_type == "evidence_list":
            receipts = _source_receipts(block, citations)
            if receipts:
                parts.append(
                    '<section class="sources" id="source-receipts">'
                    "<h2>Source receipts</h2><ol>"
                )
                for receipt in receipts:
                    parts.append(
                        f'<li id="source-{receipt.label.lower()}">'
                        f"<strong>{html.escape(receipt.label)}</strong> "
                        f"{html.escape(receipt.source)}"
                        f"<small>{html.escape(receipt.locator)}"
                        f"{' · ' if receipt.locator and receipt.detail else ''}"
                        f"{html.escape(receipt.detail)} · Evidence "
                        f"{html.escape(receipt.short_id)}</small>"
                        "</li>"
                    )
                parts.append("</ol></section>")
        elif block_type == "table":
            columns, rows = _table_data(block)
            parts.append("<table><thead><tr>")
            parts.extend(f"<th>{html.escape(str(value))}</th>" for value in columns)
            parts.append("</tr></thead><tbody>")
            for row in rows:
                parts.append("<tr>")
                parts.extend(f"<td>{html.escape(str(value))}</td>" for value in row)
                parts.append("</tr>")
            parts.append("</tbody></table>")
    parts.append("</main></body></html>")
    return "".join(parts)


def _artifact_pdf(document: dict[str, Any], metadata: ArtifactRenderMetadata) -> bytes:
    citations = _citation_context(document)
    output = io.BytesIO()
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        font_name = "STSong-Light"
    except Exception:
        font_name = "Helvetica"
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle(
        "NexusBody", parent=styles["BodyText"], fontName=font_name, leading=16, alignment=TA_LEFT
    )
    title_style = ParagraphStyle(
        "NexusTitle", parent=styles["Heading1"], fontName=font_name, leading=24
    )
    source_title_style = ParagraphStyle(
        "NexusSourceTitle", parent=styles["Heading2"], fontName=font_name, leading=20
    )
    cover_eyebrow_style = ParagraphStyle(
        "NexusCoverEyebrow",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#60716D"),
    )
    cover_title_style = ParagraphStyle(
        "NexusCoverTitle",
        parent=styles["Heading1"],
        fontName=font_name,
        fontSize=24,
        leading=28,
    )
    markdown_styles = {
        "body": body_style,
        "heading2": ParagraphStyle(
            "NexusHeading2", parent=styles["Heading2"], fontName=font_name, leading=20
        ),
        "heading3": ParagraphStyle(
            "NexusHeading3", parent=styles["Heading3"], fontName=font_name, leading=18
        ),
        "heading4": ParagraphStyle(
            "NexusHeading4", parent=styles["Heading4"], fontName=font_name, leading=16
        ),
        "list": ParagraphStyle(
            "NexusList", parent=body_style, leftIndent=16, firstLineIndent=-10, spaceAfter=4
        ),
    }
    story: list[Any] = []
    _append_pdf_delivery_cover(
        story,
        document,
        metadata,
        citations,
        cover_eyebrow_style,
        cover_title_style,
        body_style,
        font_name,
    )
    for block in _blocks(document):
        block_type = block.get("type")
        if block_type == "heading":
            story.append(Paragraph(html.escape(str(block.get("text", ""))), title_style))
        elif block_type == "paragraph":
            _append_pdf_markdown(
                story,
                str(block.get("text", "")),
                citations,
                markdown_styles,
                font_name,
            )
        elif block_type == "evidence_list":
            receipts = _source_receipts(block, citations)
            if receipts:
                story.append(Paragraph("Source receipts", source_title_style))
                for receipt in receipts:
                    parts = [
                        f"{receipt.label}. {receipt.source}",
                        receipt.locator,
                        receipt.detail,
                        f"Evidence {receipt.short_id}",
                    ]
                    receipt_text = " · ".join(part for part in parts if part)
                    story.append(Paragraph(html.escape(receipt_text), body_style))
        elif block_type == "table":
            columns, rows = _table_data(block)
            table = Table([columns, *rows], repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("FONTNAME", (0, 0), (-1, -1), font_name),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCE7E5")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#73817F")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(table)
        story.append(Spacer(1, 10))
    footer = _pdf_page_footer(document, metadata, font_name)
    SimpleDocTemplate(
        output,
        pagesize=A4,
        title=str(document.get("title", "Artifact")),
        bottomMargin=58,
    ).build(story, onFirstPage=footer, onLaterPages=footer)
    return output.getvalue()


def _artifact_csv(document: dict[str, Any]) -> bytes:
    tables = [block for block in _blocks(document) if block.get("type") == "table"]
    if not tables:
        raise ValidationError("CSV export requires at least one table block")
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    for index, block in enumerate(tables):
        columns, rows = _table_data(block)
        if index:
            writer.writerow([])
        writer.writerow(columns)
        writer.writerows(rows)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _artifact_xlsx(
    document: dict[str, Any],
    metadata: ArtifactRenderMetadata | None = None,
) -> bytes:
    tables = [block for block in _blocks(document) if block.get("type") == "table"]
    if not tables:
        raise ValidationError("XLSX export requires at least one table block")
    workbook = Workbook()
    manifest_created = _append_xlsx_delivery_manifest(
        workbook,
        document,
        metadata or ArtifactRenderMetadata(),
    )
    if not manifest_created:
        workbook.remove(workbook.active)
    for index, block in enumerate(tables, start=1):
        title = str(block.get("title") or f"Table {index}")[:31]
        sheet = workbook.create_sheet(title=title)
        columns, rows = _table_data(block)
        sheet.append(columns)
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_xlsx_delivery_manifest(
    workbook: Workbook,
    document: dict[str, Any],
    metadata: ArtifactRenderMetadata,
) -> bool:
    citations = _citation_context(document)
    items = _metadata_items(metadata, citations)
    if not items:
        return False
    sheet = workbook.active
    sheet.title = "Nexus Delivery"
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Nexus Artifact Delivery Packet"
    sheet["A2"] = _display_title(document)
    sheet["A4"] = "Field"
    sheet["B4"] = "Value"
    for row_index, (label, value) in enumerate(items, start=5):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
    sheet.freeze_panes = "A5"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 52
    sheet["A1"].font = Font(bold=True, color="173F37", size=13)
    sheet["A2"].font = Font(color="60716D", italic=True)
    header_fill = PatternFill("solid", fgColor="DCE7E5")
    for cell in sheet[4]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="173F37")
    for row in sheet.iter_rows(min_row=5, max_row=4 + len(items), max_col=2):
        row[0].font = Font(bold=True, color="60716D")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    return True


def _blocks(document: dict[str, Any]) -> list[dict[str, Any]]:
    return [item for item in document.get("blocks", []) if isinstance(item, dict)]


def _normalize_evidence_id(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip().lower()


def _unique_evidence_ids(ids: list[object]) -> list[str]:
    seen: set[str] = set()
    unique_ids: list[str] = []
    for value in ids:
        evidence_id = str(value).strip() if value is not None else ""
        key = _normalize_evidence_id(evidence_id)
        if not key or key in seen:
            continue
        seen.add(key)
        unique_ids.append(evidence_id)
    return unique_ids


def _evidence_marker_ids(text: str) -> list[str]:
    return [match.group(1) for match in EVIDENCE_MARKER_RE.finditer(text)]


def _document_evidence_ids(document: dict[str, Any]) -> list[str]:
    ids: list[object] = []
    for block in _blocks(document):
        if block.get("type") == "paragraph":
            ids.extend(_evidence_marker_ids(str(block.get("text", ""))))
        evidence_ids = block.get("evidence_revision_ids")
        if isinstance(evidence_ids, list):
            ids.extend(evidence_ids)
        if block.get("type") == "evidence_list":
            for item in block.get("items", []):
                if isinstance(item, dict):
                    ids.append(item.get("evidence_revision_id"))
    return _unique_evidence_ids(ids)


def _evidence_id_prefix(evidence_id: str) -> str:
    return evidence_id[: min(DEFAULT_ID_PREFIX_LENGTH, len(evidence_id))]


def _compact_id_part(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]", "", value)


def _compact_id_tail(evidence_id: str, prefix: str, tail_length: int) -> str:
    compact_remainder = _compact_id_part(evidence_id[len(prefix) :])
    compact_full_id = _compact_id_part(evidence_id)
    source = compact_remainder or compact_full_id or evidence_id
    return source if len(source) <= tail_length else source[-tail_length:]


def _colliding_evidence_labels(ids: list[str], prefix: str, tail_length: int) -> list[str]:
    return [f"{prefix}/{_compact_id_tail(evidence_id, prefix, tail_length)}" for evidence_id in ids]


def _short_evidence_labels(ids: list[str]) -> dict[str, str]:
    labels: dict[str, str] = {}
    groups: dict[str, list[str]] = {}
    for evidence_id in _unique_evidence_ids(ids):
        prefix = _evidence_id_prefix(evidence_id)
        groups.setdefault(prefix, []).append(evidence_id)

    for prefix, group in groups.items():
        if len(group) == 1:
            labels[_normalize_evidence_id(group[0])] = prefix
            continue
        tail_length = MIN_COLLISION_TAIL_LENGTH
        candidates = _colliding_evidence_labels(group, prefix, tail_length)
        while len(set(candidates)) < len(group) and tail_length < MAX_COLLISION_TAIL_LENGTH:
            tail_length += 2
            candidates = _colliding_evidence_labels(group, prefix, tail_length)
        collision_counts: dict[str, int] = {}
        for evidence_id, candidate in zip(group, candidates, strict=True):
            seen_count = collision_counts.get(candidate, 0)
            collision_counts[candidate] = seen_count + 1
            labels[_normalize_evidence_id(evidence_id)] = (
                f"{candidate}.{seen_count + 1}" if seen_count else candidate
            )
    return labels


def _citation_context(document: dict[str, Any]) -> _CitationContext:
    evidence_ids = _document_evidence_ids(document)
    short_labels = _short_evidence_labels(evidence_ids)
    references = {
        _normalize_evidence_id(evidence_id): _CitationReference(
            evidence_id=evidence_id,
            label=f"E{index}",
            short_id=short_labels.get(_normalize_evidence_id(evidence_id), evidence_id),
        )
        for index, evidence_id in enumerate(evidence_ids, start=1)
    }
    return _CitationContext(references=references)


def _slugify_heading(value: str, fallback: str) -> str:
    slug = re.sub(r"[^\w\u4e00-\u9fff]+", "-", value.strip().lower(), flags=re.UNICODE)
    slug = slug.strip("-")
    return slug or fallback


def _unique_anchor_id(base: str, seen: dict[str, int]) -> str:
    seen_count = seen.get(base, 0)
    seen[base] = seen_count + 1
    return f"{base}-{seen_count + 1}" if seen_count else base


def _document_outline(
    document: dict[str, Any],
    citations: _CitationContext,
) -> list[_DocumentOutlineEntry]:
    entries: list[_DocumentOutlineEntry] = []
    seen: dict[str, int] = {}
    heading_count = 0

    def add_heading(level: int, title: object) -> None:
        nonlocal heading_count
        text = str(title or "").strip()
        if not text:
            return
        heading_count += 1
        base = _slugify_heading(text, f"section-{heading_count}")
        entries.append(
            _DocumentOutlineEntry(
                anchor_id=_unique_anchor_id(base, seen),
                kind="section",
                level=max(1, min(6, level)),
                title=text,
            )
        )

    for block in _blocks(document):
        block_type = block.get("type")
        if block_type == "heading":
            add_heading(int(block.get("level", 1)), block.get("text", ""))
        elif block_type == "paragraph":
            for line in str(block.get("text", "")).splitlines():
                heading = MARKDOWN_HEADING_RE.match(line.strip())
                if heading:
                    add_heading(len(heading.group(1)), heading.group(2))
        elif block_type == "evidence_list" and _source_receipts(block, citations):
            entries.append(
                _DocumentOutlineEntry(
                    anchor_id="source-receipts",
                    kind="sources",
                    level=2,
                    title="Source receipts",
                )
            )
    return entries


def _outline_anchor_at(outline: list[_DocumentOutlineEntry], heading_index: int) -> str:
    heading_entries = [entry for entry in outline if entry.kind == "section"]
    if heading_index >= len(heading_entries):
        return ""
    return heading_entries[heading_index].anchor_id


def _html_document_map(outline: list[_DocumentOutlineEntry]) -> str:
    if len(outline) < 3:
        return ""
    parts = [
        '<nav class="delivery-map" aria-label="Document map">',
        "<h2>Document map</h2><ol>",
    ]
    for entry in outline:
        label = f"{'Source' if entry.kind == 'sources' else f'H{entry.level}'} · {entry.title}"
        parts.append(
            f'<li><a href="#{html.escape(entry.anchor_id)}">'
            f"{html.escape(label)}</a></li>"
        )
    parts.append("</ol></nav>")
    return "".join(parts)


def _display_title(document: dict[str, Any]) -> str:
    return str(document.get("title") or "Artifact")


def _display_label(value: object) -> str:
    text = str(value or "").replace("_", " ").replace("-", " ").strip()
    return text.title() if text else ""


def _short_identifier(value: str | None) -> str:
    return value[:8] if value else ""


def _format_timestamp(value: object) -> str:
    if not value:
        return ""
    if hasattr(value, "isoformat"):
        text = str(value.isoformat())  # type: ignore[attr-defined]
    else:
        text = str(value)
    return text.replace("+00:00", "Z").replace("T", " ")


def _metadata_items(
    metadata: ArtifactRenderMetadata,
    citations: _CitationContext,
) -> list[tuple[str, str]]:
    items: list[tuple[str, str]] = []
    if metadata.status:
        items.append(("Lifecycle", _display_label(metadata.status)))
    revision_parts = [
        f"v{metadata.revision_no}" if metadata.revision_no is not None else "",
        _short_identifier(metadata.revision_id),
    ]
    revision = " · ".join(part for part in revision_parts if part)
    if revision:
        items.append(("Revision", revision))
    if metadata.coverage_percent is not None:
        coverage_detail = ""
        if (
            metadata.supported_block_count is not None
            and metadata.content_block_count is not None
        ):
            coverage_detail = (
                f" · {metadata.supported_block_count}/{metadata.content_block_count} blocks"
            )
        items.append(("Coverage", f"{metadata.coverage_percent}%{coverage_detail}"))
    evidence_count = metadata.bound_evidence_count
    if evidence_count is None:
        evidence_count = len(citations.references) or None
    if evidence_count is not None:
        items.append(("Evidence", f"{evidence_count} bound references"))
    if metadata.pending_refresh_count:
        items.append(("Refresh", f"{metadata.pending_refresh_count} pending review"))
    updated_at = _format_timestamp(metadata.updated_at)
    if updated_at:
        items.append(("Updated", updated_at))
    if metadata.artifact_id:
        items.append(("Artifact ID", _short_identifier(metadata.artifact_id)))
    return items


def _markdown_delivery_metadata(
    document: dict[str, Any],
    metadata: ArtifactRenderMetadata,
    citations: _CitationContext,
) -> list[str]:
    items = _metadata_items(metadata, citations)
    if not items:
        return []
    lines = [
        "> Nexus Artifact Delivery Packet",
        f"> Title: {_display_title(document)}",
    ]
    lines.extend(f"> {label}: {value}" for label, value in items)
    return lines


def _html_delivery_cover(
    document: dict[str, Any],
    metadata: ArtifactRenderMetadata,
    citations: _CitationContext,
) -> str:
    items = _metadata_items(metadata, citations)
    if not items:
        return ""
    parts = [
        '<section class="delivery-cover" aria-label="Artifact delivery metadata">',
        '<p class="eyebrow">Nexus Artifact Delivery Packet</p>',
        f'<p class="delivery-title">{html.escape(_display_title(document))}</p>',
        '<dl class="delivery-meta">',
    ]
    for label, value in items:
        parts.append(
            "<div>"
            f"<dt>{html.escape(label)}</dt>"
            f"<dd>{html.escape(value)}</dd>"
            "</div>"
        )
    parts.append("</dl></section>")
    return "".join(parts)


def _append_pdf_delivery_cover(
    story: list[Any],
    document: dict[str, Any],
    metadata: ArtifactRenderMetadata,
    citations: _CitationContext,
    eyebrow_style: ParagraphStyle,
    title_style: ParagraphStyle,
    body_style: ParagraphStyle,
    font_name: str,
) -> None:
    items = _metadata_items(metadata, citations)
    if not items:
        return
    story.append(Paragraph("Nexus Artifact Delivery Packet", eyebrow_style))
    story.append(Paragraph(html.escape(_display_title(document)), title_style))
    table_data = [["Field", "Value"], *[[label, value] for label, value in items]]
    table = Table(table_data, hAlign="LEFT", colWidths=[110, 330])
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCE7E5")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD8D2")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D3DED9")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#173F37")),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 18))


def _pdf_footer_text(document: dict[str, Any], metadata: ArtifactRenderMetadata) -> str:
    parts = [_display_title(document)]
    if metadata.status:
        parts.append(_display_label(metadata.status))
    revision_parts = [
        f"v{metadata.revision_no}" if metadata.revision_no is not None else "",
        _short_identifier(metadata.revision_id),
    ]
    revision = " · ".join(part for part in revision_parts if part)
    if revision:
        parts.append(revision)
    return " · ".join(part for part in parts if part)


def _pdf_page_footer(document: dict[str, Any], metadata: ArtifactRenderMetadata, font_name: str):
    footer_text = _pdf_footer_text(document, metadata)

    def draw_footer(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#CBD8D2"))
        canvas.setFillColor(colors.HexColor("#60716D"))
        canvas.setFont(font_name, 7)
        width, _height = A4
        y = 28
        canvas.line(doc.leftMargin, y + 10, width - doc.rightMargin, y + 10)
        canvas.drawString(doc.leftMargin, y, footer_text)
        canvas.drawRightString(width - doc.rightMargin, y, f"Page {doc.page}")
        canvas.restoreState()

    return draw_footer


def _markdown_text_with_citations(text: str, citations: _CitationContext) -> str:
    return EVIDENCE_MARKER_RE.sub(lambda match: _citation_label(match, citations), text)


def _plain_text_with_citations(text: str, citations: _CitationContext) -> str:
    return EVIDENCE_MARKER_RE.sub(lambda match: _citation_label(match, citations), text)


def _citation_label(match: re.Match[str], citations: _CitationContext) -> str:
    reference = citations.reference(match.group(1))
    return f"[{reference.label}]" if reference else ""


def _inline_html_segment(text: str) -> str:
    rendered = html.escape(text)
    rendered = re.sub(r"`(.+?)`", r"<code>\1</code>", rendered)
    rendered = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", rendered)
    return rendered


def _html_inline_with_citations(text: str, citations: _CitationContext) -> str:
    parts: list[str] = []
    cursor = 0
    for match in EVIDENCE_MARKER_RE.finditer(text):
        parts.append(_inline_html_segment(text[cursor : match.start()]))
        reference = citations.reference(match.group(1))
        if reference:
            parts.append(
                f'<a class="citation" href="#source-{reference.label.lower()}" '
                f'title="Evidence {html.escape(reference.short_id)}">{reference.label}</a>'
            )
        cursor = match.end()
    parts.append(_inline_html_segment(text[cursor:]))
    return "".join(parts)


def _inline_pdf_segment(text: str) -> str:
    rendered = html.escape(text)
    rendered = re.sub(
        r"`(.+?)`",
        lambda match: f'<font name="Courier">{match.group(1)}</font>',
        rendered,
    )
    rendered = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", rendered)
    return rendered


def _pdf_inline_with_citations(text: str, citations: _CitationContext) -> str:
    parts: list[str] = []
    cursor = 0
    for match in EVIDENCE_MARKER_RE.finditer(text):
        parts.append(_inline_pdf_segment(text[cursor : match.start()]))
        reference = citations.reference(match.group(1))
        if reference:
            parts.append(html.escape(f"[{reference.label}]"))
        cursor = match.end()
    parts.append(_inline_pdf_segment(text[cursor:]))
    return "".join(parts)


def _plain_markdown_text_with_citations(text: str, citations: _CitationContext) -> str:
    rendered = _plain_text_with_citations(text, citations)
    rendered = re.sub(r"`(.+?)`", r"\1", rendered)
    rendered = re.sub(r"\*\*(.+?)\*\*", r"\1", rendered)
    return rendered


def _flush_html_paragraph(
    parts: list[str],
    paragraph_lines: list[str],
    citations: _CitationContext,
) -> None:
    if not paragraph_lines:
        return
    text = " ".join(line.strip() for line in paragraph_lines if line.strip())
    if text:
        parts.append(f"<p>{_html_inline_with_citations(text, citations)}</p>")
    paragraph_lines.clear()


def _is_horizontal_rule(line: str) -> bool:
    return line in {"---", "***", "___"}


def _markdown_table_cells(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def _is_markdown_table_separator(line: str) -> bool:
    cells = _markdown_table_cells(line)
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells)


def _is_markdown_table_start(lines: list[str], index: int) -> bool:
    if index + 1 >= len(lines):
        return False
    return "|" in lines[index] and _is_markdown_table_separator(lines[index + 1])


def _html_markdown_table(
    lines: list[str],
    start_index: int,
    citations: _CitationContext,
) -> tuple[str, int]:
    columns = _markdown_table_cells(lines[start_index])
    rows: list[list[str]] = []
    index = start_index + 2
    while index < len(lines) and "|" in lines[index].strip():
        rows.append(_markdown_table_cells(lines[index]))
        index += 1

    width = max([len(columns), *(len(row) for row in rows)], default=0)
    columns = (columns + [""] * width)[:width]
    rows = [(row + [""] * width)[:width] for row in rows]
    parts = ["<table><thead><tr>"]
    parts.extend(f"<th>{_html_inline_with_citations(column, citations)}</th>" for column in columns)
    parts.append("</tr></thead><tbody>")
    for row in rows:
        parts.append("<tr>")
        parts.extend(f"<td>{_html_inline_with_citations(cell, citations)}</td>" for cell in row)
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "".join(parts), index


def _html_list(
    lines: list[str],
    start_index: int,
    pattern: re.Pattern[str],
    tag_name: str,
    citations: _CitationContext,
) -> tuple[str, int]:
    parts = [f"<{tag_name}>"]
    index = start_index
    while index < len(lines):
        match = pattern.match(lines[index])
        if not match:
            break
        content = _html_inline_with_citations(match.group(match.lastindex or 1), citations)
        parts.append(f"<li>{content}</li>")
        index += 1
    parts.append(f"</{tag_name}>")
    return "".join(parts), index


def _html_markdown_blocks(
    text: str,
    citations: _CitationContext,
    outline: list[_DocumentOutlineEntry],
    heading_index: int,
) -> tuple[str, int]:
    lines = text.splitlines()
    parts: list[str] = []
    paragraph_lines: list[str] = []
    index = 0

    while index < len(lines):
        line = lines[index].rstrip()
        stripped = line.strip()
        if not stripped:
            _flush_html_paragraph(parts, paragraph_lines, citations)
            index += 1
            continue
        if _is_markdown_table_start(lines, index):
            _flush_html_paragraph(parts, paragraph_lines, citations)
            table_html, index = _html_markdown_table(lines, index, citations)
            parts.append(table_html)
            continue
        heading = MARKDOWN_HEADING_RE.match(stripped)
        if heading:
            _flush_html_paragraph(parts, paragraph_lines, citations)
            level = max(1, min(6, len(heading.group(1))))
            heading_text = _html_inline_with_citations(heading.group(2), citations)
            anchor = _outline_anchor_at(outline, heading_index)
            heading_index += 1
            anchor_attr = f' id="{html.escape(anchor)}"' if anchor else ""
            parts.append(f"<h{level}{anchor_attr}>{heading_text}</h{level}>")
            index += 1
            continue
        if _is_horizontal_rule(stripped):
            _flush_html_paragraph(parts, paragraph_lines, citations)
            parts.append("<hr>")
            index += 1
            continue
        if ORDERED_LIST_RE.match(line):
            _flush_html_paragraph(parts, paragraph_lines, citations)
            list_html, index = _html_list(lines, index, ORDERED_LIST_RE, "ol", citations)
            parts.append(list_html)
            continue
        if BULLET_LIST_RE.match(line):
            _flush_html_paragraph(parts, paragraph_lines, citations)
            list_html, index = _html_list(lines, index, BULLET_LIST_RE, "ul", citations)
            parts.append(list_html)
            continue
        paragraph_lines.append(line)
        index += 1

    _flush_html_paragraph(parts, paragraph_lines, citations)
    return "".join(parts), heading_index


def _flush_pdf_paragraph(
    story: list[Any],
    paragraph_lines: list[str],
    citations: _CitationContext,
    body_style: ParagraphStyle,
) -> None:
    if not paragraph_lines:
        return
    text = " ".join(line.strip() for line in paragraph_lines if line.strip())
    if text:
        story.append(Paragraph(_pdf_inline_with_citations(text, citations), body_style))
    paragraph_lines.clear()


def _pdf_markdown_table(
    lines: list[str],
    start_index: int,
    citations: _CitationContext,
    font_name: str,
) -> tuple[Table, int]:
    columns = _markdown_table_cells(lines[start_index])
    rows: list[list[str]] = []
    index = start_index + 2
    while index < len(lines) and "|" in lines[index].strip():
        rows.append(_markdown_table_cells(lines[index]))
        index += 1
    width = max([len(columns), *(len(row) for row in rows)], default=0)
    columns = (columns + [""] * width)[:width]
    rows = [(row + [""] * width)[:width] for row in rows]
    table_data = [
        [_plain_markdown_text_with_citations(cell, citations) for cell in row]
        for row in [columns, *rows]
    ]
    table = Table(
        table_data,
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCE7E5")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#73817F")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table, index


def _append_pdf_list(
    story: list[Any],
    lines: list[str],
    start_index: int,
    pattern: re.Pattern[str],
    ordered: bool,
    citations: _CitationContext,
    list_style: ParagraphStyle,
) -> int:
    index = start_index
    ordinal = 1
    while index < len(lines):
        match = pattern.match(lines[index])
        if not match:
            break
        marker = f"{ordinal}." if ordered else "•"
        content = _pdf_inline_with_citations(match.group(match.lastindex or 1), citations)
        story.append(Paragraph(f"{marker} {content}", list_style))
        index += 1
        ordinal += 1
    return index


def _append_pdf_markdown(
    story: list[Any],
    text: str,
    citations: _CitationContext,
    styles: dict[str, ParagraphStyle],
    font_name: str,
) -> None:
    lines = text.splitlines()
    paragraph_lines: list[str] = []
    index = 0
    while index < len(lines):
        line = lines[index].rstrip()
        stripped = line.strip()
        if not stripped:
            _flush_pdf_paragraph(story, paragraph_lines, citations, styles["body"])
            index += 1
            continue
        if _is_markdown_table_start(lines, index):
            _flush_pdf_paragraph(story, paragraph_lines, citations, styles["body"])
            table, index = _pdf_markdown_table(lines, index, citations, font_name)
            story.append(table)
            story.append(Spacer(1, 8))
            continue
        heading = MARKDOWN_HEADING_RE.match(stripped)
        if heading:
            _flush_pdf_paragraph(story, paragraph_lines, citations, styles["body"])
            level = len(heading.group(1))
            style = styles.get(f"heading{min(max(level, 2), 4)}", styles["heading4"])
            story.append(Paragraph(_pdf_inline_with_citations(heading.group(2), citations), style))
            index += 1
            continue
        if _is_horizontal_rule(stripped):
            _flush_pdf_paragraph(story, paragraph_lines, citations, styles["body"])
            story.append(Spacer(1, 10))
            index += 1
            continue
        if ORDERED_LIST_RE.match(line):
            _flush_pdf_paragraph(story, paragraph_lines, citations, styles["body"])
            index = _append_pdf_list(
                story, lines, index, ORDERED_LIST_RE, True, citations, styles["list"]
            )
            continue
        if BULLET_LIST_RE.match(line):
            _flush_pdf_paragraph(story, paragraph_lines, citations, styles["body"])
            index = _append_pdf_list(
                story, lines, index, BULLET_LIST_RE, False, citations, styles["list"]
            )
            continue
        paragraph_lines.append(line)
        index += 1
    _flush_pdf_paragraph(story, paragraph_lines, citations, styles["body"])


def _source_receipts(block: dict[str, Any], citations: _CitationContext) -> list[_SourceReceipt]:
    receipts: list[_SourceReceipt] = []
    seen: set[str] = set()
    for index, item in enumerate(block.get("items", []), start=1):
        if not isinstance(item, dict):
            continue
        evidence_id = str(item.get("evidence_revision_id") or "").strip()
        key = _normalize_evidence_id(evidence_id) or f"unbound-{index}"
        if key in seen:
            continue
        seen.add(key)
        reference = citations.reference(evidence_id)
        locator, detail = _summarize_locator(item.get("locator"))
        receipts.append(
            _SourceReceipt(
                detail=detail,
                evidence_id=evidence_id,
                label=reference.label if reference else "Unbound",
                locator=locator,
                short_id=reference.short_id if reference else "unbound",
                source=str(item.get("source") or "Evidence source"),
            )
        )
    return receipts


def _number_value(value: object) -> float | None:
    return value if isinstance(value, (int, float)) else None


def _format_bbox(value: object) -> str:
    if not isinstance(value, list) or len(value) < 4:
        return ""
    numbers = [_number_value(item) for item in value[:4]]
    if any(item is None for item in numbers):
        return ""
    return "Bounding box " + ", ".join(str(round(item or 0)) for item in numbers)


def _format_ms(value: float) -> str:
    seconds = value / 1000
    return f"{seconds:.0f}s" if value % 1000 == 0 else f"{seconds:.1f}s"


def _summarize_locator(locator: object) -> tuple[str, str]:
    if not isinstance(locator, dict):
        return "Locator", "Open exact locator"
    locator_type = str(locator.get("locator_type") or "locator")
    page_no = _number_value(locator.get("page_no"))
    char_start = _number_value(locator.get("char_start"))
    char_end = _number_value(locator.get("char_end"))
    start_ms = _number_value(locator.get("start_ms"))
    end_ms = _number_value(locator.get("end_ms"))
    sheet = str(locator.get("sheet") or "") if locator.get("sheet") else ""
    cell_range = str(locator.get("cell_range") or "") if locator.get("cell_range") else ""
    bbox = _format_bbox(locator.get("bbox"))

    if locator_type == "text_range":
        detail = (
            f"Characters {int(char_start)}-{int(char_end)}"
            if char_start is not None and char_end is not None
            else "Text span in source"
        )
        return "Text span", detail
    if locator_type == "page_region":
        detail = " · ".join(
            part for part in [f"Page {int(page_no)}" if page_no is not None else "", bbox] if part
        )
        return "Page region", detail or "Page region"
    if locator_type == "cell_range":
        return "Table cells", " · ".join(part for part in [sheet or "Sheet", cell_range] if part)
    if locator_type == "time_range":
        detail = (
            f"{_format_ms(start_ms)}-{_format_ms(end_ms)}"
            if start_ms is not None and end_ms is not None
            else "Timed media segment"
        )
        return "Media segment", detail
    return locator_type.replace("_", " ").title(), "Open exact locator"


def _table_data(block: dict[str, Any]) -> tuple[list[Any], list[list[Any]]]:
    columns = list(block.get("columns") or [])
    rows = [list(row) for row in block.get("rows", []) if isinstance(row, (list, tuple))]
    if not columns:
        width = max((len(row) for row in rows), default=0)
        columns = [f"Column {index + 1}" for index in range(width)]
    normalized = [(row + [None] * len(columns))[: len(columns)] for row in rows]
    return columns, normalized
