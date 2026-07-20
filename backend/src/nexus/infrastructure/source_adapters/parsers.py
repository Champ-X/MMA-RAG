from __future__ import annotations

import base64
import binascii
import csv
import hashlib
import io
import ipaddress
import json
import math
import mimetypes
import re
import shutil
import socket
import subprocess
import tempfile
import wave
from array import array
from dataclasses import replace
from pathlib import Path
from urllib.parse import urljoin, urlsplit

import httpx

from nexus.infrastructure.mineru import MinerURemoteAdapter
from nexus.modules.evidence.domain import EvidenceDraft, Locator
from nexus.modules.sources.domain import DerivedAsset, ParseResult, SourceVersionView
from nexus.shared.domain.enums import Modality
from nexus.shared.domain.errors import CapabilityUnavailableError, ValidationError


def _hash_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValidationError("Text source is not valid UTF-8 or GB18030")


_MARKDOWN_IMAGE_RE = re.compile(
    r"!\[(?P<alt>[^\]]*)\]\(\s*(?P<target><[^>]+>|[^\s)]+)"
    r"(?:\s+[\"'][^\"']*[\"'])?\s*\)",
    re.IGNORECASE,
)
_HTML_IMAGE_RE = re.compile(r"<img\b(?P<attrs>[^>]*)>", re.IGNORECASE)
_HTML_ATTRIBUTE_RE = re.compile(
    r"(?P<name>[\w:-]+)\s*=\s*(?P<quote>[\"'])(?P<value>.*?)(?P=quote)",
    re.IGNORECASE | re.DOTALL,
)


class ParserRouter:
    """Format router with explicit capability failure instead of semantic fallback."""

    DOCUMENT_EXTENSIONS = {".pdf", ".doc", ".docx", ".ppt", ".pptx"}

    def __init__(
        self,
        *,
        mineru: MinerURemoteAdapter | None = None,
        media_analyzer: object | None = None,
        remote_image_timeout_seconds: float = 12.0,
        remote_image_max_bytes: int = 10 * 1024 * 1024,
        remote_image_max_count: int = 30,
        allow_private_networks: bool = False,
    ) -> None:
        self.mineru = mineru
        self.media_analyzer = media_analyzer
        self.remote_image_timeout_seconds = remote_image_timeout_seconds
        self.remote_image_max_bytes = remote_image_max_bytes
        self.remote_image_max_count = remote_image_max_count
        self.allow_private_networks = allow_private_networks

    def parse(
        self,
        *,
        content: bytes,
        filename: str,
        mime_type: str,
        source_version: SourceVersionView,
    ) -> ParseResult:
        extension = Path(filename).suffix.lower()
        if extension in self.DOCUMENT_EXTENSIONS:
            return self._parse_mineru(content, filename, source_version)
        if extension in {".md", ".markdown"} or mime_type == "text/markdown":
            return self._parse_markdown(content, filename, source_version)
        if extension == ".txt" or mime_type.startswith("text/plain"):
            return self._coerce(
                self._parse_text(content, filename, source_version, markdown=False)
            )
        if extension == ".csv" or mime_type == "text/csv":
            return self._coerce(self._parse_csv(content, filename, source_version))
        if extension in {".xlsx", ".xlsm"}:
            return self._coerce(self._parse_xlsx(content, filename, source_version))
        if extension == ".xls":
            return self._coerce(self._parse_xls(content, filename, source_version))
        if source_version.modality == Modality.IMAGE:
            return self._coerce(self._parse_image(content, filename, source_version))
        if source_version.modality == Modality.AUDIO:
            return self._coerce(self._parse_audio(content, filename, source_version))
        if source_version.modality == Modality.VIDEO:
            return self._parse_video(content, filename, source_version)
        raise CapabilityUnavailableError(
            "No parser is configured for this source format",
            details={"filename": filename, "mime_type": mime_type},
        )

    @staticmethod
    def _coerce(
        parsed: tuple[list[EvidenceDraft], dict[str, object], dict[str, str]],
    ) -> ParseResult:
        drafts, manifest, capabilities = parsed
        return ParseResult(drafts=tuple(drafts), manifest=manifest, capabilities=capabilities)

    def _parse_text(
        self,
        content: bytes,
        filename: str,
        source_version: SourceVersionView,
        *,
        markdown: bool,
    ) -> tuple[list[EvidenceDraft], dict[str, object], dict[str, str]]:
        text = _decode_text(content).replace("\r\n", "\n")
        sections: list[tuple[str, int, int]] = []
        if markdown:
            matches = list(re.finditer(r"(?m)^#{1,6}\s+.+$", text))
            if matches:
                if matches[0].start() > 0 and text[: matches[0].start()].strip():
                    sections.append((text[: matches[0].start()].strip(), 0, matches[0].start()))
                for index, match in enumerate(matches):
                    end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
                    sections.append((text[match.start() : end].strip(), match.start(), end))
        if not sections:
            cursor = 0
            for part in re.split(r"\n\s*\n", text):
                clean = part.strip()
                if not clean:
                    continue
                start = text.find(clean, cursor)
                end = start + len(clean)
                cursor = end
                sections.append((clean, start, end))
        drafts: list[EvidenceDraft] = []
        for section, start, _end in sections:
            for offset in range(0, len(section), 1600):
                chunk = section[offset : offset + 1600].strip()
                if not chunk:
                    continue
                chunk_start = start + offset
                chunk_end = chunk_start + len(chunk)
                anchor = f"chars:{chunk_start}-{chunk_end}"
                drafts.append(
                    EvidenceDraft(
                        unit_type="section" if markdown else "paragraph",
                        native_anchor=anchor,
                        fingerprint=_hash_text(chunk),
                        modality=Modality.TEXT,
                        evidence_type="markdown_section" if markdown else "text_paragraph",
                        text_content=chunk,
                        searchable_text=chunk,
                        content_hash=_hash_text(chunk),
                        locator=Locator(
                            locator_type="text_range",
                            char_start=chunk_start,
                            char_end=chunk_end,
                            extra={"filename": filename},
                        ),
                        ordinal=len(drafts),
                        provenance={"parser": "native-text-v1"},
                    )
                )
        if not drafts:
            raise ValidationError("Text source contains no searchable content")
        return (
            drafts,
            {"parser": "native-text-v1", "source_version_id": source_version.id},
            {"parse_structure": "ready", "text_index": "pending"},
        )

    def _parse_markdown(
        self,
        content: bytes,
        filename: str,
        source_version: SourceVersionView,
    ) -> ParseResult:
        text = _decode_text(content).replace("\r\n", "\n").replace("\r", "\n")
        references = _markdown_image_references(text)[: self.remote_image_max_count]
        safe_text = _mask_markdown_data_images(text, references)
        text_drafts: list[EvidenceDraft] = []
        text_manifest: dict[str, object] = {
            "parser": "native-markdown-multimodal-v2",
            "source_version_id": source_version.id,
        }
        try:
            text_drafts, _manifest, _capabilities = self._parse_text(
                safe_text.encode("utf-8"),
                filename,
                source_version,
                markdown=True,
            )
        except ValidationError:
            # Image-only Markdown is still a valid multimodal source.
            if not references:
                raise

        image_drafts: list[EvidenceDraft] = []
        assets: dict[str, DerivedAsset] = {}
        failures: list[dict[str, object]] = []
        for reference in references:
            target = str(reference["target"])
            try:
                image_data, content_type, resolved_target = self._materialize_markdown_image(
                    target
                )
                content_type, width, height, image_format = _inspect_image(
                    image_data,
                    content_type=content_type,
                )
            except (CapabilityUnavailableError, ValidationError) as exc:
                failures.append(
                    {
                        "reference": _safe_image_reference(target),
                        "error": exc.code,
                    }
                )
                continue

            digest = hashlib.sha256(image_data).hexdigest()
            suffix = mimetypes.guess_extension(content_type) or _image_suffix(image_format)
            object_key = f"derived/{source_version.id}/markdown/{digest}{suffix}"
            assets.setdefault(
                object_key,
                DerivedAsset(
                    object_key=object_key,
                    data=image_data,
                    content_type=content_type,
                    content_hash=digest,
                    role="document_image",
                    source_path=resolved_target,
                ),
            )
            alt = str(reference.get("alt") or "").strip()
            caption: str | None = None
            caption_failed = False
            if self.media_analyzer is not None and getattr(
                self.media_analyzer, "image_configured", False
            ):
                try:
                    caption = self.media_analyzer.caption_image(  # type: ignore[attr-defined]
                        image_data,
                        mime_type=content_type,
                        task_role="document_figure_caption",
                    )
                except Exception:
                    caption_failed = True
            description_parts = [
                f"Embedded image from {filename}",
                *([f"Alt text: {alt}"] if alt else []),
                *(
                    [f"Dimensions: {width} by {height} pixels"]
                    if width is not None and height is not None
                    else []
                ),
                *([f"Visual description: {caption}"] if caption else []),
            ]
            description = "\n".join(description_parts)
            reference_start = int(reference["start"])
            reference_end = int(reference["end"])
            image_drafts.append(
                EvidenceDraft(
                    unit_type="image",
                    native_anchor=f"chars:{reference_start}-{reference_end}:image",
                    fingerprint=_hash_text(f"{digest}:{reference_start}:{description}"),
                    modality=Modality.IMAGE,
                    evidence_type="markdown_image",
                    text_content=description,
                    searchable_text=f"{filename} {alt} {caption or ''}".strip(),
                    content_hash=_hash_text(f"{digest}:{description}"),
                    locator=Locator(
                        locator_type="text_range",
                        char_start=reference_start,
                        char_end=reference_end,
                        extra={
                            "filename": filename,
                            "object_key": object_key,
                            "image_reference": _safe_image_reference(target),
                            "resolved_reference": _safe_image_reference(resolved_target),
                            "width": width,
                            "height": height,
                            "scope": "embedded_image",
                        },
                    ),
                    ordinal=len(text_drafts) + len(image_drafts),
                    quality_flags=(
                        ("caption_unavailable",)
                        if caption_failed or not caption
                        else ()
                    ),
                    provenance={
                        "parser": "native-markdown-multimodal-v2",
                        "format": image_format,
                        "source_kind": _markdown_image_source_kind(target),
                    },
                )
            )

        if not text_drafts and not image_drafts:
            raise ValidationError("Markdown source contains no publishable content")
        extracted = len(image_drafts)
        referenced = len(references)
        text_manifest.update(
            {
                "embedded_image_references": referenced,
                "extracted_images": extracted,
                "image_failures": failures,
                "image_reference_limit": self.remote_image_max_count,
            }
        )
        return ParseResult(
            drafts=tuple([*text_drafts, *image_drafts]),
            manifest=text_manifest,
            capabilities={
                "parse_structure": "ready",
                "text_index": "pending" if text_drafts else "disabled",
                "extracted_images": (
                    "ready"
                    if referenced and extracted == referenced
                    else "partial"
                    if extracted
                    else "failed"
                    if referenced
                    else "disabled"
                ),
                "image_pipeline": "pending" if extracted else "disabled",
            },
            derived_assets=tuple(assets.values()),
        )

    def _materialize_markdown_image(self, target: str) -> tuple[bytes, str, str]:
        value = target.strip().removeprefix("<").removesuffix(">")
        if value.lower().startswith("data:"):
            header, separator, payload = value.partition(",")
            if not separator or ";base64" not in header.lower():
                raise ValidationError("Markdown image data URI must use base64 encoding")
            content_type = header[5:].split(";", 1)[0].strip().lower() or "image/png"
            try:
                data = base64.b64decode(payload, validate=True)
            except (ValueError, binascii.Error) as exc:
                raise ValidationError("Markdown image data URI is not valid base64") from exc
            if len(data) > self.remote_image_max_bytes:
                raise ValidationError("Markdown image exceeds the configured byte limit")
            return data, content_type, "inline:data-uri"
        if value.startswith(("http://", "https://")):
            return self._download_remote_image(value)
        raise ValidationError(
            "Relative Markdown image is unavailable outside a connector asset bundle",
            details={"reference": _safe_image_reference(value)},
        )

    def _download_remote_image(self, initial_url: str) -> tuple[bytes, str, str]:
        url = initial_url
        headers = {
            "User-Agent": "MMA-RAG-Nexus/2.0",
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
        }
        try:
            with httpx.Client(
                timeout=self.remote_image_timeout_seconds,
                follow_redirects=False,
            ) as client:
                for _ in range(6):
                    self._validate_remote_image_url(url)
                    with client.stream("GET", url, headers=headers) as response:
                        if response.status_code in {301, 302, 303, 307, 308}:
                            location = response.headers.get("location")
                            if not location:
                                raise ValidationError("Image redirect has no Location header")
                            url = urljoin(url, location)
                            continue
                        response.raise_for_status()
                        data = bytearray()
                        for chunk in response.iter_bytes():
                            data.extend(chunk)
                            if len(data) > self.remote_image_max_bytes:
                                raise ValidationError(
                                    "Markdown image exceeds the configured byte limit"
                                )
                        content_type = (
                            response.headers.get("content-type", "application/octet-stream")
                            .split(";", 1)[0]
                            .strip()
                            .lower()
                        )
                        return bytes(data), content_type, str(response.url)
        except httpx.HTTPError as exc:
            raise CapabilityUnavailableError(
                "Markdown image could not be downloaded",
                details={
                    "error_type": type(exc).__name__,
                    "hostname": urlsplit(url).hostname,
                },
            ) from exc
        raise ValidationError("Markdown image exceeded the redirect limit")

    def _validate_remote_image_url(self, value: str) -> None:
        parsed = urlsplit(value)
        if (
            parsed.scheme not in {"http", "https"}
            or not parsed.hostname
            or parsed.username
            or parsed.password
        ):
            raise ValidationError("Markdown image URL is not an allowed absolute URL")
        try:
            addresses = {
                item[4][0]
                for item in socket.getaddrinfo(
                    parsed.hostname,
                    parsed.port,
                    type=socket.SOCK_STREAM,
                )
            }
        except OSError as exc:
            raise CapabilityUnavailableError(
                "Markdown image hostname could not be resolved",
                details={"error_type": type(exc).__name__},
            ) from exc
        if not addresses:
            raise ValidationError("Markdown image hostname did not resolve")
        if self.allow_private_networks:
            return
        if any(not ipaddress.ip_address(address).is_global for address in addresses):
            raise ValidationError(
                "Markdown image resolves to a non-public network",
                details={"host": parsed.hostname},
            )

    def _parse_csv(
        self, content: bytes, filename: str, source_version: SourceVersionView
    ) -> tuple[list[EvidenceDraft], dict[str, object], dict[str, str]]:
        text = _decode_text(content)
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            raise ValidationError("CSV source is empty")
        header = rows[0]
        drafts: list[EvidenceDraft] = []
        summary = (
            f"Table {filename}; columns: {', '.join(header)}; data rows: {max(0, len(rows) - 1)}"
        )
        drafts.append(
            EvidenceDraft(
                unit_type="sheet_summary",
                native_anchor="sheet:CSV:summary",
                fingerprint=_hash_text(summary),
                modality=Modality.TABLE,
                evidence_type="sheet_summary",
                text_content=summary,
                searchable_text=summary,
                content_hash=_hash_text(summary),
                locator=Locator(
                    locator_type="cell_range",
                    sheet="CSV",
                    cell_range=f"A1:{_column(len(header))}{len(rows)}",
                ),
                provenance={"parser": "native-csv-v1"},
            )
        )
        drafts.extend(
            _table_column_profile_drafts(
                filename=filename,
                sheet="CSV",
                headers=header,
                rows=rows[1:],
                first_data_row=2,
                parser="native-csv-v2",
                ordinal_start=len(drafts),
            )
        )
        for start in range(1, len(rows), 25):
            group = rows[start : start + 25]
            rendered = "\n".join([" | ".join(header)] + [" | ".join(row) for row in group])
            end_column = _column(max(len(row) for row in [header, *group]))
            cell_range = f"A{start + 1}:{end_column}{start + len(group)}"
            drafts.append(
                EvidenceDraft(
                    unit_type="row_block",
                    native_anchor=f"sheet:CSV:{cell_range}",
                    fingerprint=_hash_text(rendered),
                    modality=Modality.TABLE,
                    evidence_type="table_row_block",
                    text_content=rendered,
                    searchable_text=rendered,
                    content_hash=_hash_text(rendered),
                    locator=Locator(locator_type="cell_range", sheet="CSV", cell_range=cell_range),
                    ordinal=len(drafts),
                    provenance={"parser": "native-csv-v1"},
                )
            )
        return (
            drafts,
            {"parser": "native-csv-v1", "rows": len(rows)},
            {
                "parse_structure": "ready",
                "table_structure": "ready",
                "text_index": "pending",
            },
        )

    def _parse_xlsx(
        self, content: bytes, filename: str, source_version: SourceVersionView
    ) -> tuple[list[EvidenceDraft], dict[str, object], dict[str, str]]:
        try:
            import openpyxl
        except ImportError as exc:
            raise CapabilityUnavailableError("openpyxl is required for XLSX ingestion") from exc
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False, read_only=False)
        drafts: list[EvidenceDraft] = []
        for sheet in workbook.worksheets:
            max_row = max(sheet.max_row, 1)
            max_column = max(sheet.max_column, 1)
            headers = [
                str(sheet.cell(1, column).value or "") for column in range(1, max_column + 1)
            ]
            summary = (
                f"Workbook {filename}; sheet {sheet.title}; "
                f"columns: {', '.join(headers)}; rows: {max_row}"
            )
            drafts.append(
                EvidenceDraft(
                    unit_type="sheet_summary",
                    native_anchor=f"sheet:{sheet.title}:summary",
                    fingerprint=_hash_text(summary),
                    modality=Modality.TABLE,
                    evidence_type="sheet_summary",
                    text_content=summary,
                    searchable_text=summary,
                    content_hash=_hash_text(summary),
                    locator=Locator(
                        locator_type="cell_range",
                        sheet=sheet.title,
                        cell_range=f"A1:{_column(max_column)}{max_row}",
                    ),
                    ordinal=len(drafts),
                    provenance={"parser": "native-xlsx-v1"},
                )
            )
            table_rows = [
                [
                    str(sheet.cell(row_no, column).value or "")
                    for column in range(1, max_column + 1)
                ]
                for row_no in range(2, max_row + 1)
            ]
            drafts.extend(
                _table_column_profile_drafts(
                    filename=filename,
                    sheet=sheet.title,
                    headers=headers,
                    rows=table_rows,
                    first_data_row=2,
                    parser="native-xlsx-v2",
                    ordinal_start=len(drafts),
                )
            )
            for start in range(2, max_row + 1, 25):
                end = min(max_row, start + 24)
                lines = [" | ".join(headers)]
                for row_no in range(start, end + 1):
                    values = [
                        str(sheet.cell(row_no, column).value or "")
                        for column in range(1, max_column + 1)
                    ]
                    lines.append(" | ".join(values))
                rendered = "\n".join(lines)
                cell_range = f"A{start}:{_column(max_column)}{end}"
                drafts.append(
                    EvidenceDraft(
                        unit_type="row_block",
                        native_anchor=f"sheet:{sheet.title}:{cell_range}",
                        fingerprint=_hash_text(rendered),
                        modality=Modality.TABLE,
                        evidence_type="table_row_block",
                        text_content=rendered,
                        searchable_text=rendered,
                        content_hash=_hash_text(rendered),
                        locator=Locator(
                            locator_type="cell_range", sheet=sheet.title, cell_range=cell_range
                        ),
                        ordinal=len(drafts),
                        provenance={"parser": "native-xlsx-v1"},
                    )
                )
        if not drafts:
            raise ValidationError("XLSX source contains no worksheets")
        return (
            drafts,
            {"parser": "native-xlsx-v1", "sheets": workbook.sheetnames},
            {
                "parse_structure": "ready",
                "table_structure": "ready",
                "text_index": "pending",
            },
        )

    def _parse_xls(
        self, content: bytes, filename: str, source_version: SourceVersionView
    ) -> tuple[list[EvidenceDraft], dict[str, object], dict[str, str]]:
        try:
            import xlrd
        except ImportError as exc:
            raise CapabilityUnavailableError("xlrd is required for XLS ingestion") from exc
        try:
            workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        except Exception as exc:
            raise ValidationError("XLS source cannot be decoded") from exc
        drafts: list[EvidenceDraft] = []
        sheet_names = workbook.sheet_names()
        for sheet_name in sheet_names:
            sheet = workbook.sheet_by_name(sheet_name)
            max_row = max(int(sheet.nrows), 1)
            max_column = max(int(sheet.ncols), 1)
            headers = [
                _xls_cell_text(sheet.cell_value(0, column)) if sheet.nrows else ""
                for column in range(max_column)
            ]
            summary = (
                f"Workbook {filename}; sheet {sheet.name}; "
                f"columns: {', '.join(headers)}; rows: {max_row}"
            )
            drafts.append(
                EvidenceDraft(
                    unit_type="sheet_summary",
                    native_anchor=f"sheet:{sheet.name}:summary",
                    fingerprint=_hash_text(summary),
                    modality=Modality.TABLE,
                    evidence_type="sheet_summary",
                    text_content=summary,
                    searchable_text=summary,
                    content_hash=_hash_text(summary),
                    locator=Locator(
                        locator_type="cell_range",
                        sheet=sheet.name,
                        cell_range=f"A1:{_column(max_column)}{max_row}",
                    ),
                    ordinal=len(drafts),
                    provenance={"parser": "native-xls-v1"},
                )
            )
            table_rows = [
                [
                    _xls_cell_text(sheet.cell_value(row_no, column))
                    for column in range(max_column)
                ]
                for row_no in range(1, int(sheet.nrows))
            ]
            drafts.extend(
                _table_column_profile_drafts(
                    filename=filename,
                    sheet=sheet.name,
                    headers=headers,
                    rows=table_rows,
                    first_data_row=2,
                    parser="native-xls-v2",
                    ordinal_start=len(drafts),
                )
            )
            for start in range(1, int(sheet.nrows), 25):
                end = min(int(sheet.nrows), start + 25)
                lines = [" | ".join(headers)]
                for row_no in range(start, end):
                    values = [
                        _xls_cell_text(sheet.cell_value(row_no, column))
                        for column in range(max_column)
                    ]
                    lines.append(" | ".join(values))
                rendered = "\n".join(lines)
                cell_range = f"A{start + 1}:{_column(max_column)}{end}"
                drafts.append(
                    EvidenceDraft(
                        unit_type="row_block",
                        native_anchor=f"sheet:{sheet.name}:{cell_range}",
                        fingerprint=_hash_text(rendered),
                        modality=Modality.TABLE,
                        evidence_type="table_row_block",
                        text_content=rendered,
                        searchable_text=rendered,
                        content_hash=_hash_text(rendered),
                        locator=Locator(
                            locator_type="cell_range",
                            sheet=sheet.name,
                            cell_range=cell_range,
                        ),
                        ordinal=len(drafts),
                        provenance={"parser": "native-xls-v1"},
                    )
                )
        workbook.release_resources()
        if not drafts:
            raise ValidationError("XLS source contains no worksheets")
        return (
            drafts,
            {"parser": "native-xls-v1", "sheets": sheet_names},
            {
                "parse_structure": "ready",
                "table_structure": "ready",
                "text_index": "pending",
            },
        )

    def _parse_image(
        self, content: bytes, filename: str, source_version: SourceVersionView
    ) -> tuple[list[EvidenceDraft], dict[str, object], dict[str, str]]:
        try:
            from PIL import Image

            image = Image.open(io.BytesIO(content))
            image.load()
            width, height = image.size
            image_format = image.format
        except Exception as exc:
            raise ValidationError("Image source cannot be decoded") from exc
        caption: str | None = None
        caption_status = "not_configured"
        ocr_text: str | None = None
        ocr_status = "not_configured"
        if self.media_analyzer is not None and getattr(
            self.media_analyzer, "image_configured", False
        ):
            try:
                caption = self.media_analyzer.caption_image(  # type: ignore[attr-defined]
                    content,
                    mime_type=Image.MIME.get(image_format, "image/png"),
                    task_role="image_caption",
                )
                caption_status = "ready"
            except Exception:
                caption_status = "failed"
        if self.media_analyzer is not None and getattr(
            self.media_analyzer, "ocr_configured", False
        ):
            try:
                ocr_text = self.media_analyzer.ocr_image(content)  # type: ignore[attr-defined]
                ocr_status = "ready"
            except Exception:
                ocr_status = "failed"
        enrichments = []
        if caption:
            enrichments.append(f"Visual description: {caption}")
        if ocr_text:
            enrichments.append(f"Visible text and layout (OCR):\n{ocr_text}")
        description = "\n".join([f"Image {filename}, {width} by {height} pixels", *enrichments])
        quality_flags: list[str] = []
        if not ocr_text:
            quality_flags.append("ocr_unavailable")
        if not caption:
            quality_flags.append("caption_unavailable")
        drafts = [
            EvidenceDraft(
                unit_type="image",
                native_anchor="image:whole",
                fingerprint=source_version.content_hash,
                modality=Modality.IMAGE,
                evidence_type="whole_image",
                text_content=description,
                searchable_text=f"{filename} {description}",
                content_hash=_hash_text(description),
                locator=Locator(
                    locator_type="image",
                    extra={
                        "width": width,
                        "height": height,
                        "filename": filename,
                        "object_key": source_version.object_key,
                        "scope": "whole_image",
                    },
                ),
                ordinal=0,
                quality_flags=tuple(quality_flags),
                provenance={
                    "parser": "native-image-structure-v3",
                    "format": image_format,
                    "caption_model": (
                        getattr(self.media_analyzer, "image_model", None) if caption else None
                    ),
                },
            )
        ]
        return (
            drafts,
            {
                "parser": "native-image-structure-v3",
                "width": width,
                "height": height,
            },
            {
                "parse_structure": "ready",
                "global_visual": "ready",
                "caption": caption_status,
                "ocr": ocr_status,
                "visual_index": "pending",
            },
        )

    def _parse_audio(
        self, content: bytes, filename: str, source_version: SourceVersionView
    ) -> tuple[list[EvidenceDraft], dict[str, object], dict[str, str]]:
        samples, sample_rate = _decode_audio(content, filename)
        duration_ms = int(len(samples) / max(sample_rate, 1) * 1000)
        window_samples = max(sample_rate * 30, 1)
        overlap_samples = min(sample_rate * 2, max(window_samples - 1, 0))
        step_samples = max(window_samples - overlap_samples, 1)
        asr_configured = bool(
            self.media_analyzer is not None
            and getattr(self.media_analyzer, "audio_configured", False)
        )
        asr_ready = 0
        asr_failed = 0
        drafts: list[EvidenceDraft] = []
        for start_sample in range(0, max(len(samples), 1), step_samples):
            if start_sample and len(samples) - start_sample <= overlap_samples:
                break
            segment = samples[start_sample : start_sample + window_samples]
            if not segment and drafts:
                break
            start_ms = int(start_sample / max(sample_rate, 1) * 1000)
            end_ms = min(
                duration_ms,
                int((start_sample + len(segment)) / max(sample_rate, 1) * 1000),
            )
            transcript: str | None = None
            if asr_configured and segment:
                try:
                    transcript = self.media_analyzer.transcribe_audio(  # type: ignore[attr-defined]
                        _encode_wav(segment, sample_rate),
                        filename=f"{Path(filename).stem}-{start_ms}-{end_ms}.wav",
                        task_role="audio_transcription",
                    )
                    asr_ready += 1
                except Exception:
                    asr_failed += 1
            locator_description = (
                f"Audio segment from {filename}, {start_ms / 1000:.2f} to "
                f"{end_ms / 1000:.2f} seconds"
            )
            description = (
                f"{locator_description}\nTranscript: {transcript}"
                if transcript
                else locator_description
            )
            acoustic_vector = _acoustic_vector(segment)
            speech_ratio = _speech_activity_ratio(segment, sample_rate)
            drafts.append(
                EvidenceDraft(
                    unit_type="audio_segment",
                    native_anchor=f"time:{start_ms}-{end_ms}",
                    fingerprint=_hash_text(f"{source_version.content_hash}:{start_ms}:{end_ms}"),
                    modality=Modality.AUDIO,
                    evidence_type="audio_segment",
                    text_content=description,
                    searchable_text=f"{filename} {description}",
                    content_hash=_hash_text(description),
                    locator=Locator(
                        locator_type="time_range",
                        start_ms=start_ms,
                        end_ms=end_ms,
                        extra={"filename": filename, "object_key": source_version.object_key},
                    ),
                    ordinal=len(drafts),
                    quality_flags=(
                        ("speaker_unavailable",)
                        if transcript
                        else ("transcript_unavailable", "speaker_unavailable")
                    ),
                    provenance={
                        "parser": "segmented-audio-evidence-v2",
                        "sample_rate": sample_rate,
                        "segment_overlap_ms": int(overlap_samples / sample_rate * 1000),
                        "asr_scope": "segment",
                        "vad": "energy-adaptive-v1",
                        "speech_activity_ratio": speech_ratio,
                        "acoustic_encoder": "pcm-envelope-zcr-v1",
                        "acoustic_vector": acoustic_vector,
                    },
                )
            )
        return (
            drafts,
            {
                "parser": "segmented-audio-evidence-v2",
                "duration_ms": duration_ms,
                "sample_rate": sample_rate,
                "segments": len(drafts),
                "segment_window_ms": 30_000,
                "segment_overlap_ms": int(overlap_samples / sample_rate * 1000),
                "transcribed_segments": asr_ready,
            },
            {
                "parse_structure": "ready",
                "segmentation": "ready",
                "vad": "ready",
                "asr": (
                    "ready"
                    if asr_configured and asr_ready == len(drafts)
                    else "partial"
                    if asr_ready
                    else "failed"
                    if asr_configured and asr_failed
                    else "not_configured"
                ),
                "diarization": "not_configured",
                "topic": "not_configured",
                "acoustic_index": "pending",
            },
        )

    def _parse_video(
        self, content: bytes, filename: str, source_version: SourceVersionView
    ) -> ParseResult:
        duration_ms: int | None = None
        ffprobe = shutil.which("ffprobe")
        ffmpeg = shutil.which("ffmpeg")
        frames: list[tuple[bytes, int]] = []
        audio_track: bytes | None = None
        has_audio = False
        strategy = "metadata_only"
        frame_interval_ms = 10_000
        if ffprobe and ffmpeg:
            suffix = Path(filename).suffix[:10] or ".bin"
            with tempfile.TemporaryDirectory(prefix="nexus-video-") as directory:
                input_path = Path(directory) / f"source{suffix}"
                input_path.write_bytes(content)
                result = subprocess.run(
                    [
                        ffprobe,
                        "-v",
                        "error",
                        "-show_entries",
                        "format=duration:stream=codec_type",
                        "-of",
                        "json",
                        str(input_path),
                    ],
                    capture_output=True,
                    text=True,
                    timeout=30,
                    check=False,
                )
                if result.returncode == 0:
                    payload = json.loads(result.stdout or "{}")
                    duration_ms = int(float(payload.get("format", {}).get("duration", 0)) * 1000)
                    has_audio = any(
                        stream.get("codec_type") == "audio"
                        for stream in payload.get("streams", [])
                        if isinstance(stream, dict)
                    )
                is_long = bool(duration_ms and duration_ms > 120_000)
                strategy = "long_overlapping_chapters" if is_long else "short_semantic_timeline"
                frame_interval_ms = 30_000 if is_long else 10_000
                frame_pattern = Path(directory) / "frame-%04d.png"
                extracted = subprocess.run(
                    [
                        ffmpeg,
                        "-v",
                        "error",
                        "-i",
                        str(input_path),
                        "-vf",
                        (
                            f"fps=fps=1/{frame_interval_ms / 1000:g}:"
                            "start_time=0:round=up,"
                            "scale=1280:-2:force_original_aspect_ratio=decrease"
                        ),
                        str(frame_pattern),
                    ],
                    capture_output=True,
                    timeout=120,
                    check=False,
                )
                if extracted.returncode == 0:
                    frames = [
                        (path.read_bytes(), index * frame_interval_ms)
                        for index, path in enumerate(sorted(Path(directory).glob("frame-*.png")))
                    ]
                if has_audio:
                    audio_result = subprocess.run(
                        [
                            ffmpeg,
                            "-v",
                            "error",
                            "-i",
                            str(input_path),
                            "-vn",
                            "-ac",
                            "1",
                            "-ar",
                            "16000",
                            "-f",
                            "wav",
                            "pipe:1",
                        ],
                        capture_output=True,
                        timeout=300,
                        check=False,
                    )
                    if audio_result.returncode == 0 and audio_result.stdout:
                        audio_track = audio_result.stdout
        assets: list[DerivedAsset] = []
        scene_drafts: list[EvidenceDraft] = []
        audio_drafts: list[EvidenceDraft] = []
        audio_transcripts: list[tuple[int, int, str]] = []
        caption_ready = 0
        caption_failed = 0
        chapter_caption_ready = 0
        chapter_caption_failed = 0
        audio_asr_ready = 0
        audio_asr_failed = 0

        audio_object_key: str | None = None
        if audio_track:
            audio_digest = hashlib.sha256(audio_track).hexdigest()
            audio_object_key = f"derived/{source_version.id}/video/{audio_digest}.wav"
            assets.append(
                DerivedAsset(
                    object_key=audio_object_key,
                    data=audio_track,
                    content_type="audio/wav",
                    content_hash=audio_digest,
                    role="video_audio_track",
                    source_path=f"{Path(filename).stem}-audio.wav",
                )
            )
            audio_samples, audio_rate = _decode_audio(audio_track, "video-audio.wav")
            window = max(audio_rate * 30, 1)
            overlap = min(audio_rate * 2, max(window - 1, 0))
            step = max(window - overlap, 1)
            asr_configured = bool(
                self.media_analyzer is not None
                and getattr(self.media_analyzer, "audio_configured", False)
            )
            for start_sample in range(0, max(len(audio_samples), 1), step):
                if start_sample and len(audio_samples) - start_sample <= overlap:
                    break
                segment = audio_samples[start_sample : start_sample + window]
                if not segment and audio_drafts:
                    break
                start_ms = int(start_sample / max(audio_rate, 1) * 1000)
                end_ms = int((start_sample + len(segment)) / max(audio_rate, 1) * 1000)
                transcript: str | None = None
                if asr_configured and segment:
                    try:
                        transcript = self.media_analyzer.transcribe_audio(  # type: ignore[attr-defined]
                            _encode_wav(segment, audio_rate),
                            filename=f"video-audio-{start_ms}-{end_ms}.wav",
                            task_role="video_audio_transcription",
                        )
                        audio_asr_ready += 1
                        audio_transcripts.append((start_ms, end_ms, transcript))
                    except Exception:
                        audio_asr_failed += 1
                base = (
                    f"Audio track from video {filename}, {start_ms / 1000:.2f} to "
                    f"{end_ms / 1000:.2f} seconds"
                )
                description = f"{base}\nTranscript: {transcript}" if transcript else base
                audio_drafts.append(
                    EvidenceDraft(
                        unit_type="video_audio_segment",
                        native_anchor=f"audio-time:{start_ms}-{end_ms}",
                        fingerprint=_hash_text(
                            f"{source_version.content_hash}:audio:{start_ms}:{end_ms}"
                        ),
                        modality=Modality.AUDIO,
                        evidence_type="video_audio_segment",
                        text_content=description,
                        searchable_text=f"{filename} {description}",
                        content_hash=_hash_text(description),
                        locator=Locator(
                            locator_type="time_range",
                            start_ms=start_ms,
                            end_ms=end_ms,
                            extra={
                                "filename": filename,
                                "video_object_key": source_version.object_key,
                                "audio_object_key": audio_object_key,
                            },
                        ),
                        ordinal=len(frames) + len(audio_drafts),
                        quality_flags=(
                            ("speaker_unavailable",)
                            if transcript
                            else ("transcript_unavailable", "speaker_unavailable")
                        ),
                        provenance={
                            "parser": "video-audio-shared-pipeline-v2",
                            "sample_rate": audio_rate,
                            "segment_overlap_ms": int(overlap / audio_rate * 1000),
                            "vad": "energy-adaptive-v1",
                            "speech_activity_ratio": _speech_activity_ratio(segment, audio_rate),
                            "acoustic_vector": _acoustic_vector(segment),
                            "acoustic_encoder": "pcm-envelope-zcr-v1",
                        },
                    )
                )

        frame_captions: list[str | None] = []
        visual_vectors: list[list[float]] = []
        for index, (frame, _start_ms) in enumerate(frames):
            digest = hashlib.sha256(frame).hexdigest()
            object_key = f"derived/{source_version.id}/video/{digest}.png"
            assets.append(
                DerivedAsset(
                    object_key=object_key,
                    data=frame,
                    content_type="image/png",
                    content_hash=digest,
                    role="video_keyframe",
                    source_path=f"frame-{index + 1:04d}.png",
                )
            )
            try:
                from PIL import Image

                keyframe = Image.open(io.BytesIO(frame))
                keyframe.load()
                visual_vectors.append(_visual_vector(keyframe))
            except Exception as exc:
                raise ValidationError("Extracted video keyframe cannot be decoded") from exc
            frame_caption: str | None = None
            if self.media_analyzer is not None and getattr(
                self.media_analyzer, "image_configured", False
            ):
                try:
                    frame_caption = self.media_analyzer.caption_image(  # type: ignore[attr-defined]
                        frame,
                        mime_type="image/png",
                        task_role="video_understanding",
                    )
                    caption_ready += 1
                except Exception:
                    caption_failed += 1
            frame_captions.append(frame_caption)

        chapter_summaries: list[tuple[int, int, str]] = []
        if (
            frames
            and self.media_analyzer is not None
            and getattr(self.media_analyzer, "image_configured", False)
            and hasattr(self.media_analyzer, "caption_image_sequence")
        ):
            chapter_window_ms = (
                480_000
                if strategy.startswith("long")
                else max(
                    duration_ms or frame_interval_ms,
                    frame_interval_ms,
                )
            )
            chapter_step_ms = 470_000 if strategy.startswith("long") else chapter_window_ms
            chapter_start = 0
            while chapter_start < max(duration_ms or frame_interval_ms, frame_interval_ms):
                chapter_end = min(
                    duration_ms or chapter_start + chapter_window_ms,
                    chapter_start + chapter_window_ms,
                )
                chapter_frames = [
                    frame
                    for frame, timestamp_ms in frames
                    if chapter_start <= timestamp_ms < chapter_end
                ]
                if chapter_frames:
                    try:
                        summary = self.media_analyzer.caption_image_sequence(  # type: ignore[attr-defined]
                            chapter_frames,
                            task_role="video_understanding",
                        )
                        chapter_summaries.append((chapter_start, chapter_end, summary))
                        chapter_caption_ready += 1
                    except Exception:
                        chapter_caption_failed += 1
                if not strategy.startswith("long"):
                    break
                chapter_start += chapter_step_ms

        for index, (frame, start_ms) in enumerate(frames):
            digest = hashlib.sha256(frame).hexdigest()
            object_key = f"derived/{source_version.id}/video/{digest}.png"
            end_ms = min(
                duration_ms or start_ms + frame_interval_ms,
                start_ms + frame_interval_ms,
            )
            locator_description = (
                f"Video scene {index + 1} from {filename}, "
                f"{start_ms / 1000:.2f} to {end_ms / 1000:.2f} seconds"
            )
            frame_caption = frame_captions[index]
            scene_transcripts = [
                transcript
                for audio_start, audio_end, transcript in audio_transcripts
                if audio_start < end_ms and audio_end > start_ms
            ]
            temporal_context: list[str] = []
            if index > 0 and frame_captions[index - 1]:
                temporal_context.append(f"Previous keyframe: {frame_captions[index - 1]}")
            if frame_caption:
                temporal_context.append(f"Current keyframe: {frame_caption}")
            aligned_chapter_summaries = [
                summary
                for chapter_start, chapter_end, summary in chapter_summaries
                if chapter_start < end_ms and chapter_end > start_ms
            ]
            if aligned_chapter_summaries:
                temporal_context.append(f"Temporal chapter: {' '.join(aligned_chapter_summaries)}")
            if scene_transcripts:
                temporal_context.append(f"Aligned audio: {' '.join(scene_transcripts)}")
            description = (
                f"{locator_description}\n" + "\n".join(temporal_context)
                if temporal_context
                else locator_description
            )
            scene_drafts.append(
                EvidenceDraft(
                    unit_type="video_scene",
                    native_anchor=f"time:{start_ms}-{end_ms}",
                    fingerprint=_hash_text(f"{source_version.content_hash}:{start_ms}:{digest}"),
                    modality=Modality.VIDEO,
                    evidence_type="video_scene",
                    text_content=description,
                    searchable_text=f"{filename} {description}",
                    content_hash=_hash_text(description),
                    locator=Locator(
                        locator_type="time_range",
                        start_ms=start_ms,
                        end_ms=end_ms,
                        extra={"filename": filename, "keyframe_object_key": object_key},
                    ),
                    ordinal=index,
                    quality_flags=(
                        ("action_inferred_from_keyframes",)
                        if frame_caption and scene_transcripts
                        else ("transcript_unavailable", "action_inferred_from_keyframes")
                        if frame_caption
                        else (
                            "frame_caption_unavailable",
                            *(() if scene_transcripts else ("transcript_unavailable",)),
                            "action_recognition_unavailable",
                        )
                    ),
                    provenance={
                        "parser": "ffmpeg-adaptive-multimodal-video-v2",
                        "strategy": strategy,
                        "chapter_window_ms": (
                            480_000 if strategy.startswith("long") else duration_ms
                        ),
                        "chapter_overlap_ms": 10_000 if strategy.startswith("long") else 0,
                        "frame_visual_encoder": "luminance-histogram-v1",
                        "frame_visual_vector": visual_vectors[index],
                        "temporal_chapter_aligned": bool(aligned_chapter_summaries),
                        "audio_aligned": bool(scene_transcripts),
                    },
                )
            )
        if not scene_drafts:
            description = f"Video {filename}"
            if duration_ms:
                description += f", duration {duration_ms / 1000:.2f} seconds"
            scene_drafts.append(
                EvidenceDraft(
                    unit_type="video_scene",
                    native_anchor=f"time:0-{duration_ms or 0}",
                    fingerprint=source_version.content_hash,
                    modality=Modality.VIDEO,
                    evidence_type="video_scene",
                    text_content=description,
                    searchable_text=f"{filename} {description}",
                    content_hash=_hash_text(description),
                    locator=Locator(
                        locator_type="time_range",
                        start_ms=0,
                        end_ms=duration_ms,
                        extra={"filename": filename},
                    ),
                    quality_flags=tuple(
                        ["scene_analysis_unavailable", "keyframe_unavailable"]
                        + ([] if audio_transcripts else ["transcript_unavailable"])
                    ),
                    provenance={"parser": "video-metadata-fallback-v1"},
                )
            )
        drafts = [*scene_drafts, *audio_drafts]
        return ParseResult(
            drafts=tuple(drafts),
            manifest={
                "parser": (
                    "ffmpeg-adaptive-multimodal-video-v2"
                    if frames
                    else "video-metadata-fallback-v1"
                ),
                "strategy": strategy,
                "duration_ms": duration_ms,
                "scenes": len(scene_drafts),
                "keyframes": len(frames),
                "frame_captions": caption_ready,
                "chapter_captions": chapter_caption_ready,
                "has_audio": has_audio,
                "audio_segments": len(audio_drafts),
                "audio_transcripts": audio_asr_ready,
                "chapter_windows": (
                    math.ceil((duration_ms or 0) / 470_000) if strategy.startswith("long") else 1
                ),
            },
            capabilities={
                "parse_structure": "ready",
                "scene": "ready" if frames else "failed",
                "keyframe": "ready" if frames else "failed",
                "frame_visual_index": "pending" if frames else "failed",
                "frame_caption": (
                    "ready"
                    if frames and caption_ready == len(frames)
                    else "partial"
                    if caption_ready
                    else "failed"
                    if caption_failed or frames
                    else "not_configured"
                ),
                "temporal_chapter": (
                    "ready"
                    if chapter_summaries and not chapter_caption_failed
                    else "partial"
                    if chapter_summaries
                    else "failed"
                    if chapter_caption_failed
                    else "not_configured"
                ),
                "audio_track": "ready" if audio_track else "failed" if has_audio else "disabled",
                "audio_vad": "ready" if audio_track else "disabled",
                "transcript": (
                    "ready"
                    if audio_drafts and audio_asr_ready == len(audio_drafts)
                    else "partial"
                    if audio_asr_ready
                    else "failed"
                    if audio_asr_failed
                    else "not_configured"
                    if audio_track
                    else "disabled"
                ),
                "ocr": "not_configured",
                "action": "derived" if caption_ready > 1 else "not_configured",
            },
            derived_assets=tuple(assets),
        )

    def _parse_mineru(
        self, content: bytes, filename: str, source_version: SourceVersionView
    ) -> ParseResult:
        if self.mineru is None:
            raise CapabilityUnavailableError(
                "MinerU Precision API adapter is not configured",
                details={"capability": "mineru_precision_api", "filename": filename},
            )
        extraction = self.mineru.extract(content=content, filename=filename)
        image_assets: dict[str, DerivedAsset] = {}
        image_keys: dict[str, str] = {}
        image_source_paths: dict[str, str] = {}
        for image in extraction.images:
            digest = hashlib.sha256(image.data).hexdigest()
            suffix = Path(image.name).suffix.lower() or ".bin"
            object_key = f"derived/{source_version.id}/mineru/{digest}{suffix}"
            asset = DerivedAsset(
                object_key=object_key,
                data=image.data,
                content_type=mimetypes.guess_type(image.name)[0] or "application/octet-stream",
                content_hash=digest,
                role="document_image",
                source_path=image.path,
            )
            image_assets[object_key] = asset
            image_source_paths[object_key] = image.path
            for candidate in (image.path, image.name, Path(image.path).name):
                normalized = _normalize_mineru_asset_path(candidate)
                if normalized:
                    image_keys[normalized] = object_key

        drafts: list[EvidenceDraft] = []
        figure_contexts: list[tuple[int, int, str, str]] = []
        referenced_object_keys: set[str] = set()
        for ordinal, unit in enumerate(extraction.content_list):
            unit_type = str(unit.get("type") or "document_block").lower()
            page_no = _mineru_page_no(unit)
            bbox = _mineru_bbox(unit.get("bbox"))
            image_path = _mineru_image_path(unit)
            normalized_image_path = _normalize_mineru_asset_path(image_path)
            object_key = image_keys.get(normalized_image_path) or image_keys.get(
                _normalize_mineru_asset_path(Path(normalized_image_path).name)
            )
            if object_key:
                referenced_object_keys.add(object_key)
            text = _mineru_text(unit, filename=filename, ordinal=ordinal)
            if object_key and unit_type != "table":
                content_hash = image_assets[object_key].content_hash
                modality = Modality.IMAGE
                evidence_type = (
                    "document_image" if unit_type == "image" else f"document_{unit_type}"
                )
                generic = f"Extracted document {unit_type} {ordinal + 1} from {filename}"
                if not text:
                    text = generic
                if self.media_analyzer is not None and getattr(
                    self.media_analyzer, "image_configured", False
                ):
                    try:
                        caption = self.media_analyzer.caption_image(  # type: ignore[attr-defined]
                            image_assets[object_key].data,
                            mime_type=image_assets[object_key].content_type,
                            task_role="document_figure_caption",
                        )
                        text = caption if text == generic else f"{text}\n{caption}"
                        content_hash = _hash_text(f"{image_assets[object_key].content_hash}:{text}")
                    except Exception:
                        pass
            elif unit_type == "table":
                content_hash = _hash_text(text)
                modality = Modality.TABLE
                evidence_type = "document_table"
            else:
                content_hash = _hash_text(text)
                modality = Modality.TEXT
                evidence_type = unit_type
            if not text and not object_key:
                continue
            anchor = f"page:{page_no}:{unit_type}:{ordinal}"
            locator_extra: dict[str, object] = {"filename": filename}
            if image_path:
                locator_extra["mineru_image_path"] = image_path
            if object_key:
                locator_extra["object_key"] = object_key
            provenance: dict[str, object] = {
                "parser": "mineru-precision-api-v1",
                "mineru_task_id": extraction.task_id,
                "mineru_model": self.mineru.model,
            }
            quality_flags: tuple[str, ...] = ()
            if (
                modality == Modality.IMAGE
                and object_key
                and text.startswith("Extracted document ")
            ):
                quality_flags = ("caption_unavailable",)
            drafts.append(
                EvidenceDraft(
                    unit_type=unit_type,
                    native_anchor=anchor,
                    fingerprint=content_hash,
                    modality=modality,
                    evidence_type=evidence_type,
                    text_content=text,
                    searchable_text=text,
                    content_hash=content_hash,
                    locator=Locator(
                        locator_type="page_region",
                        page_no=page_no,
                        bbox=bbox,
                        extra=locator_extra,
                    ),
                    ordinal=ordinal,
                    quality_flags=quality_flags,
                    provenance=provenance,
                )
            )
            if (
                modality == Modality.IMAGE
                and object_key
                and text
                and not text.startswith("Extracted document ")
            ):
                figure_contexts.append((page_no, ordinal, text, object_key))

        # MinerU result archives may contain valid extracted figures that are referenced
        # only from Markdown, or whose content-list path differs by a leading directory.
        # Keeping the immutable asset without publishing Evidence made those figures
        # invisible to retrieval and to the material inspector. Publish every remaining
        # parser-owned image explicitly; provenance records that its page anchor was not
        # available instead of silently pretending a precise location.
        next_ordinal = len(extraction.content_list)
        for offset, (object_key, asset) in enumerate(image_assets.items()):
            if object_key in referenced_object_keys:
                continue
            source_path = image_source_paths.get(object_key, asset.source_path or "")
            text = f"Extracted document image from {filename}"
            quality_flags: tuple[str, ...] = ("parser_asset_without_content_anchor",)
            if self.media_analyzer is not None and getattr(
                self.media_analyzer, "image_configured", False
            ):
                try:
                    text = self.media_analyzer.caption_image(  # type: ignore[attr-defined]
                        asset.data,
                        mime_type=asset.content_type,
                        task_role="document_figure_caption",
                    )
                    quality_flags = ("parser_asset_without_content_anchor",)
                except Exception:
                    quality_flags = (
                        "parser_asset_without_content_anchor",
                        "caption_unavailable",
                    )
            ordinal = next_ordinal + offset
            drafts.append(
                EvidenceDraft(
                    unit_type="image",
                    native_anchor=f"asset:{asset.content_hash}",
                    fingerprint=_hash_text(f"{asset.content_hash}:{text}"),
                    modality=Modality.IMAGE,
                    evidence_type="document_image",
                    text_content=text,
                    searchable_text=text,
                    content_hash=_hash_text(f"{asset.content_hash}:{text}"),
                    locator=Locator(
                        locator_type="document_asset",
                        extra={
                            "filename": filename,
                            "mineru_image_path": source_path,
                            "object_key": object_key,
                            "anchor_precision": "document",
                        },
                    ),
                    ordinal=ordinal,
                    quality_flags=quality_flags,
                    provenance={
                        "parser": "mineru-precision-api-v1",
                        "mineru_task_id": extraction.task_id,
                        "mineru_model": self.mineru.model,
                        "asset_recovery": "unmatched_zip_image",
                    },
                )
            )
        captions_written_back = 0
        for page_no, figure_ordinal, caption, object_key in figure_contexts:
            candidates = [
                (index, draft)
                for index, draft in enumerate(drafts)
                if draft.modality == Modality.TEXT and draft.locator.page_no == page_no
            ]
            if not candidates:
                candidates = [
                    (index, draft)
                    for index, draft in enumerate(drafts)
                    if draft.modality == Modality.TEXT
                ]
            if not candidates:
                continue
            target_index, target = min(
                candidates, key=lambda item: abs(item[1].ordinal - figure_ordinal)
            )
            figure_block = f"\n\n[Figure evidence · {object_key}]\n{caption}"
            enriched = f"{target.text_content}{figure_block}"
            related_object_keys = sorted(
                set(
                    [
                        *target.locator.extra.get("related_image_object_keys", []),
                        object_key,
                    ]
                )
            )
            drafts[target_index] = replace(
                target,
                fingerprint=_hash_text(enriched),
                text_content=enriched,
                searchable_text=enriched,
                content_hash=_hash_text(enriched),
                locator=replace(
                    target.locator,
                    extra={
                        **target.locator.extra,
                        # A text citation can now materialize its nearest figure inline.
                        "object_key": target.locator.extra.get("object_key") or object_key,
                        "related_image_object_keys": related_object_keys,
                    },
                ),
                provenance={
                    **target.provenance,
                    "document_figure_writeback": True,
                    "document_figure_object_keys": sorted(
                        set(
                            [
                                *target.provenance.get("document_figure_object_keys", []),
                                object_key,
                            ]
                        )
                    ),
                },
            )
            captions_written_back += 1
        if not drafts and extraction.markdown.strip():
            fallback, _manifest, _capabilities = self._parse_text(
                extraction.markdown.encode("utf-8"),
                filename,
                source_version,
                markdown=True,
            )
            drafts = [
                replace(
                    draft,
                    provenance={
                        "parser": "mineru-precision-api-markdown-v1",
                        "mineru_task_id": extraction.task_id,
                        "mineru_model": self.mineru.model,
                    },
                )
                for draft in fallback
            ]
        if not drafts:
            raise ValidationError("MinerU produced no citable evidence blocks")
        return ParseResult(
            drafts=tuple(drafts),
            manifest={
                "parser": "mineru-precision-api-v1",
                "task_id": extraction.task_id,
                "model": self.mineru.model,
                "content_units": len(extraction.content_list),
                "extracted_images": len(image_assets),
                "unanchored_images_recovered": len(
                    set(image_assets) - referenced_object_keys
                ),
                "figure_captions_written_back": captions_written_back,
            },
            capabilities={
                "parse_structure": "ready",
                "text_index": "pending",
                "extracted_images": "ready" if image_assets else "disabled",
                "image_pipeline": "pending" if image_assets else "disabled",
                "figure_caption_writeback": (
                    "ready"
                    if captions_written_back == len(figure_contexts) and figure_contexts
                    else "partial"
                    if captions_written_back
                    else "disabled"
                    if not figure_contexts
                    else "failed"
                ),
                "table_structure": "ready",
            },
            derived_assets=tuple(image_assets.values()),
        )


def _markdown_image_references(text: str) -> list[dict[str, object]]:
    references: list[dict[str, object]] = []
    for match in _MARKDOWN_IMAGE_RE.finditer(text):
        references.append(
            {
                "alt": match.group("alt"),
                "target": match.group("target"),
                "start": match.start(),
                "end": match.end(),
                "target_start": match.start("target"),
                "target_end": match.end("target"),
            }
        )
    for match in _HTML_IMAGE_RE.finditer(text):
        attributes: dict[str, tuple[str, int, int]] = {}
        attrs_text = match.group("attrs")
        attrs_offset = match.start("attrs")
        for attribute in _HTML_ATTRIBUTE_RE.finditer(attrs_text):
            attributes[attribute.group("name").lower()] = (
                attribute.group("value"),
                attrs_offset + attribute.start("value"),
                attrs_offset + attribute.end("value"),
            )
        source = attributes.get("src")
        if source is None:
            continue
        references.append(
            {
                "alt": attributes.get("alt", ("", 0, 0))[0],
                "target": source[0],
                "start": match.start(),
                "end": match.end(),
                "target_start": source[1],
                "target_end": source[2],
            }
        )
    return sorted(references, key=lambda item: (int(item["start"]), int(item["end"])))


def _mask_markdown_data_images(text: str, references: list[dict[str, object]]) -> str:
    masked = text
    for reference in reversed(references):
        target = str(reference["target"])
        if not target.lower().lstrip("<").startswith("data:"):
            continue
        start = int(reference["target_start"])
        end = int(reference["target_end"])
        replacement = "inline:image".ljust(end - start)[: end - start]
        masked = f"{masked[:start]}{replacement}{masked[end:]}"
    return masked


def _inspect_image(
    content: bytes,
    *,
    content_type: str,
) -> tuple[str, int | None, int | None, str]:
    normalized_type = content_type.split(";", 1)[0].strip().lower()
    if normalized_type == "image/svg+xml" or re.match(
        rb"^\s*(?:<\?xml[^>]*>\s*)?<svg\b",
        content[:2048],
        re.IGNORECASE,
    ):
        return "image/svg+xml", None, None, "SVG"
    try:
        from PIL import Image

        with Image.open(io.BytesIO(content)) as image:
            image.load()
            width, height = image.size
            image_format = image.format or "IMAGE"
            detected_type = Image.MIME.get(image_format, normalized_type)
    except Exception as exc:
        raise ValidationError("Markdown image cannot be decoded") from exc
    if not detected_type.startswith("image/"):
        raise ValidationError("Markdown image response is not an image")
    return detected_type, width, height, image_format


def _image_suffix(image_format: str) -> str:
    return {
        "JPEG": ".jpg",
        "PNG": ".png",
        "GIF": ".gif",
        "WEBP": ".webp",
        "BMP": ".bmp",
        "TIFF": ".tiff",
        "AVIF": ".avif",
        "SVG": ".svg",
    }.get(image_format.upper(), ".img")


def _safe_image_reference(value: str) -> str:
    clean = value.strip().removeprefix("<").removesuffix(">")
    if clean.lower().startswith("data:"):
        return "inline:data-uri"
    parsed = urlsplit(clean)
    if parsed.scheme in {"http", "https"} and parsed.hostname:
        port = f":{parsed.port}" if parsed.port else ""
        return f"{parsed.scheme}://{parsed.hostname}{port}{parsed.path}"
    return clean[:512]


def _markdown_image_source_kind(value: str) -> str:
    clean = value.strip().removeprefix("<").lower()
    if clean.startswith("data:"):
        return "inline_base64"
    if clean.startswith(("http://", "https://")):
        return "remote_url"
    return "relative_reference"


def _xls_cell_text(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "")


def _table_column_profile_drafts(
    *,
    filename: str,
    sheet: str,
    headers: list[str],
    rows: list[list[str]],
    first_data_row: int,
    parser: str,
    ordinal_start: int,
) -> list[EvidenceDraft]:
    """Publish searchable column semantics retained by the legacy table pipeline."""

    drafts: list[EvidenceDraft] = []
    for column_index, raw_name in enumerate(headers):
        name = raw_name.strip() or f"Column {_column(column_index + 1)}"
        values = [
            str(row[column_index]).strip() if column_index < len(row) else "" for row in rows
        ]
        non_empty = [value for value in values if value]
        unique_values = list(dict.fromkeys(non_empty))
        numeric_values: list[float] = []
        for value in non_empty:
            try:
                parsed_number = float(value)
                if math.isfinite(parsed_number):
                    numeric_values.append(parsed_number)
            except ValueError:
                continue
        if non_empty and len(numeric_values) / len(non_empty) >= 0.8:
            dtype = "numeric"
        elif unique_values and len(unique_values) <= max(20, int(len(non_empty) * 0.1)):
            dtype = "categorical"
        else:
            dtype = "text"

        examples = [
            value if len(value) <= 60 else f"{value[:57]}..." for value in unique_values[:5]
        ]
        details = [
            f"Column profile for {filename}; sheet {sheet}; column {name}",
            f"type: {dtype}",
            f"non-empty values: {len(non_empty)}",
            f"missing values: {len(values) - len(non_empty)}",
            f"unique values: {len(unique_values)}",
        ]
        if examples:
            details.append(f"examples: {', '.join(examples)}")
        if dtype == "numeric" and numeric_values:
            details.append(
                "numeric statistics: "
                f"min {_format_table_number(min(numeric_values))}, "
                f"max {_format_table_number(max(numeric_values))}, "
                f"mean {_format_table_number(sum(numeric_values) / len(numeric_values))}, "
                f"count {len(numeric_values)}"
            )
        rendered = "; ".join(details)
        column = _column(column_index + 1)
        last_data_row = max(first_data_row, first_data_row + len(rows) - 1)
        drafts.append(
            EvidenceDraft(
                unit_type="column_profile",
                native_anchor=f"sheet:{sheet}:column:{column}",
                fingerprint=_hash_text(rendered),
                modality=Modality.TABLE,
                evidence_type="table_column_profile",
                text_content=rendered,
                searchable_text=rendered,
                content_hash=_hash_text(rendered),
                locator=Locator(
                    locator_type="cell_range",
                    sheet=sheet,
                    cell_range=f"{column}{first_data_row}:{column}{last_data_row}",
                    extra={
                        "column_name": name,
                        "data_type": dtype,
                        "non_empty_count": len(non_empty),
                        "missing_count": len(values) - len(non_empty),
                        "unique_count": len(unique_values),
                        "examples": examples,
                    },
                ),
                ordinal=ordinal_start + len(drafts),
                provenance={"parser": parser},
            )
        )
    return drafts


def _format_table_number(value: float) -> str:
    if math.isfinite(value) and abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.4g}"


def _column(value: int) -> str:
    value = max(value, 1)
    result = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _mineru_page_no(unit: dict[str, object]) -> int:
    raw = unit.get("page_idx")
    if raw is not None:
        try:
            return max(1, int(raw) + 1)
        except (TypeError, ValueError):
            return 1
    raw = unit.get("page_no") or unit.get("page") or 1
    try:
        return max(1, int(raw))
    except (TypeError, ValueError):
        return 1


def _normalize_mineru_asset_path(value: object) -> str:
    """Normalize paths from MinerU content-list and result ZIP without trusting them."""

    if not isinstance(value, str):
        return ""
    normalized = value.strip().replace("\\", "/")
    while normalized.startswith("./"):
        normalized = normalized[2:]
    return normalized.lstrip("/")


def _mineru_image_path(unit: dict[str, object]) -> str:
    """Read the image reference used by current and older MinerU result schemas."""

    for key in ("img_path", "image_path", "image", "path"):
        value = unit.get(key)
        if isinstance(value, str) and value.strip():
            return value
        if isinstance(value, dict):
            for nested_key in ("path", "img_path", "image_path", "src"):
                nested = value.get(nested_key)
                if isinstance(nested, str) and nested.strip():
                    return nested
    return ""


def _mineru_bbox(raw: object) -> tuple[float, float, float, float] | None:
    if not isinstance(raw, (list, tuple)) or len(raw) < 4:
        return None
    try:
        return tuple(float(value) for value in raw[:4])  # type: ignore[return-value]
    except (TypeError, ValueError):
        return None


def _mineru_text(unit: dict[str, object], *, filename: str, ordinal: int) -> str:
    values: list[str] = []
    for key in ("text", "content", "table_body"):
        value = unit.get(key)
        if isinstance(value, str) and value.strip():
            values.append(value.strip())
    for key in ("image_caption", "image_footnote", "table_caption", "table_footnote"):
        value = unit.get(key)
        if isinstance(value, str) and value.strip():
            values.append(value.strip())
        elif isinstance(value, list):
            values.extend(str(item).strip() for item in value if str(item).strip())
    if values:
        return "\n".join(dict.fromkeys(values))
    if str(unit.get("type") or "").lower() == "image":
        return f"Extracted document image {ordinal + 1} from {filename}"
    return ""


def _visual_vector(image: object) -> list[float]:
    grayscale = image.convert("L").resize((64, 64))  # type: ignore[attr-defined]
    histogram = [float(value) for value in grayscale.histogram()[:256]]
    norm = math.sqrt(sum(value * value for value in histogram)) or 1.0
    return [value / norm for value in histogram]


def _decode_audio(content: bytes, filename: str) -> tuple[list[int], int]:
    if Path(filename).suffix.lower() == ".wav":
        try:
            with wave.open(io.BytesIO(content), "rb") as reader:
                if reader.getsampwidth() == 2:
                    channels = reader.getnchannels()
                    sample_rate = reader.getframerate()
                    raw = reader.readframes(reader.getnframes())
                    values = array("h")
                    values.frombytes(raw[: len(raw) - (len(raw) % 2)])
                    if channels > 1:
                        mono = [
                            int(sum(values[index : index + channels]) / max(channels, 1))
                            for index in range(0, len(values), channels)
                        ]
                    else:
                        mono = list(values)
                    return mono, sample_rate
        except wave.Error:
            pass
    ffmpeg = shutil.which("ffmpeg")
    if ffmpeg:
        decoded = subprocess.run(
            [
                ffmpeg,
                "-v",
                "error",
                "-i",
                "pipe:0",
                "-ac",
                "1",
                "-ar",
                "16000",
                "-f",
                "s16le",
                "pipe:1",
            ],
            input=content,
            capture_output=True,
            timeout=120,
            check=False,
        )
        if decoded.returncode == 0 and decoded.stdout:
            values = array("h")
            values.frombytes(decoded.stdout[: len(decoded.stdout) - (len(decoded.stdout) % 2)])
            return list(values), 16000
    raise ValidationError(
        "Audio source cannot be decoded into PCM",
        details={"filename": filename, "ffmpeg_configured": bool(ffmpeg)},
    )


def _encode_wav(samples: list[int], sample_rate: int) -> bytes:
    buffer = io.BytesIO()
    values = array("h", samples)
    with wave.open(buffer, "wb") as output:
        output.setnchannels(1)
        output.setsampwidth(2)
        output.setframerate(sample_rate)
        output.writeframes(values.tobytes())
    return buffer.getvalue()


def _acoustic_vector(samples: list[int]) -> list[float]:
    if not samples:
        return [1.0] + [0.0] * 255
    envelope: list[float] = []
    crossings: list[float] = []
    for index in range(128):
        start = index * len(samples) // 128
        end = max(start + 1, (index + 1) * len(samples) // 128)
        window = samples[start:end]
        rms = math.sqrt(sum(float(value) ** 2 for value in window) / max(len(window), 1))
        envelope.append(rms / 32768.0)
        crossing_count = sum(
            1
            for left, right in zip(window, window[1:], strict=False)
            if (left < 0 <= right) or (right < 0 <= left)
        )
        crossings.append(crossing_count / max(len(window) - 1, 1))
    vector = envelope + crossings
    norm = math.sqrt(sum(value * value for value in vector))
    if norm == 0:
        vector[0] = 1.0
        return vector
    return [value / norm for value in vector]


def _speech_activity_ratio(samples: list[int], sample_rate: int) -> float:
    if not samples:
        return 0.0
    frame = max(1, int(sample_rate * 0.02))
    energies = [
        math.sqrt(
            sum(float(value) ** 2 for value in samples[start : start + frame])
            / max(len(samples[start : start + frame]), 1)
        )
        for start in range(0, len(samples), frame)
    ]
    sorted_energy = sorted(energies)
    noise_floor = sorted_energy[max(0, len(sorted_energy) // 5 - 1)] if sorted_energy else 0.0
    threshold = max(180.0, noise_floor * 2.6)
    return round(sum(energy >= threshold for energy in energies) / max(len(energies), 1), 4)
