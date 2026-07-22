from __future__ import annotations

import io
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from nexus.bootstrap import NexusContainer, build_container
from nexus.config import NexusSettings
from nexus.infrastructure.celery.tasks import _sync_scheduled_news
from nexus.infrastructure.operations.service import restore_backup_to_empty
from nexus.infrastructure.postgres.models import Space
from nexus.modules.artifacts.renderers import ArtifactRenderMetadata, render_artifact
from nexus.shared.domain.errors import ConflictError


def test_backup_restore_drill_to_empty_root(tmp_path: Path, nexus: NexusContainer) -> None:
    space = nexus.spaces.create(name="Restore Drill", slug="restore-drill")
    source = nexus.ingestion.ingest_bytes(
        space_id=space.id,
        filename="restore.md",
        content=b"# Restore\n\nThe authoritative raw object survives a restore drill.",
        mime_type="text/markdown",
    )
    backup = nexus.operations.backup(tmp_path / "backups")
    target_db = tmp_path / "restored" / "nexus.db"
    target_blobs = tmp_path / "restored" / "blobs"
    report = restore_backup_to_empty(
        Path(backup["path"]),
        target_database_url=f"sqlite:///{target_db}",
        target_blob_root=target_blobs,
    )
    assert report["readiness"]["search"] == "rebuild_required"

    restored = build_container(
        NexusSettings(
            environment="test",
            database_url=f"sqlite:///{target_db}",
            auto_create_schema=False,
            blob_backend="filesystem",
            blob_root=target_blobs,
            qdrant_url=None,
            redis_url=None,
            agent_runtime="native",
        )
    )
    try:
        with restored.database.transaction() as session:
            assert session.scalar(select(func.count(Space.id))) == 1
        assert restored.blob_store.get(source.source_version.object_key).startswith(b"# Restore")
    finally:
        restored.database.engine.dispose()

    with pytest.raises(ConflictError):
        restore_backup_to_empty(
            Path(backup["path"]),
            target_database_url=f"sqlite:///{target_db}",
            target_blob_root=target_blobs,
        )


def test_canonical_table_csv_xlsx_pdf_exports() -> None:
    document = {
        "schema": "nexus.block-document.v1",
        "title": "Budget Evidence",
        "blocks": [
            {"type": "heading", "level": 1, "text": "预算证据"},
            {"type": "paragraph", "text": "All values remain in the canonical revision."},
            {
                "type": "table",
                "title": "Budget",
                "columns": ["Item", "USD"],
                "rows": [["Launch", 250000], ["Reserve", 50000]],
            },
        ],
    }
    json_bytes, _, _ = render_artifact(document, "json")
    assert b"nexus.block-document.v1" in json_bytes
    csv_bytes, _, _ = render_artifact(document, "csv")
    assert "Launch,250000" in csv_bytes.decode("utf-8-sig")
    metadata = ArtifactRenderMetadata(
        artifact_id="artifact-budget-123456",
        bound_evidence_count=1,
        content_block_count=1,
        coverage_percent=100,
        revision_id="revision-budget-123456",
        revision_no=3,
        status="published",
        supported_block_count=1,
        updated_at="2026-07-21T11:00:00Z",
    )
    xlsx_bytes, _, _ = render_artifact(document, "xlsx", metadata)
    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=False)
    assert workbook.sheetnames[:2] == ["Nexus Delivery", "Budget"]
    assert workbook["Nexus Delivery"]["A1"].value == "Nexus Artifact Delivery Packet"
    assert workbook["Nexus Delivery"]["B5"].value == "Published"
    assert workbook["Nexus Delivery"]["B6"].value == "v3 · revision"
    assert workbook["Nexus Delivery"]["B7"].value == "100% · 1/1 blocks"
    assert workbook["Budget"]["B2"].value == 250000
    pdf_bytes, _, _ = render_artifact(document, "pdf")
    assert pdf_bytes.startswith(b"%PDF")


def test_artifact_delivery_exports_render_auditable_citations() -> None:
    first_id = "019f7e45-42a7-777e-9a7a-b348139f6c41"
    second_id = "019f7e45-42a7-777e-9a7a-b348139f6c99"
    metadata = ArtifactRenderMetadata(
        artifact_id="artifact-alpha-123456",
        artifact_type="research_report",
        bound_evidence_count=2,
        content_block_count=5,
        coverage_percent=80,
        pending_refresh_count=1,
        revision_id="revision-alpha-123456",
        revision_no=7,
        status="candidate",
        supported_block_count=4,
        updated_at="2026-07-21T10:30:00Z",
    )
    document = {
        "schema": "nexus.block-document.v1",
        "title": "Citation Delivery",
        "blocks": [
            {"type": "heading", "level": 1, "text": "Citation Delivery"},
            {
                "type": "paragraph",
                "text": (
                    f"First grounded claim[evidence:{first_id}] and "
                    f"second grounded claim[evidence:{second_id}]. "
                    f"Repeat the first source[evidence:{first_id}]."
                ),
                "evidence_revision_ids": [second_id, first_id],
            },
            {
                "type": "evidence_list",
                "items": [
                    {
                        "evidence_revision_id": first_id,
                        "source": "architecture.md",
                        "locator": {
                            "locator_type": "text_range",
                            "char_start": 10,
                            "char_end": 28,
                        },
                    },
                    {
                        "evidence_revision_id": first_id,
                        "source": "duplicate.md",
                    },
                    {
                        "evidence_revision_id": second_id,
                        "source": "paper.pdf",
                        "locator": {
                            "locator_type": "page_region",
                            "page_no": 5,
                            "bbox": [121, 83, 878, 414],
                        },
                    },
                ],
            },
        ],
    }

    json_bytes, _, _ = render_artifact(document, "json", metadata)
    json_text = json_bytes.decode("utf-8")
    assert "[evidence:" in json_text
    assert "Nexus Artifact Delivery Packet" not in json_text

    markdown_bytes, _, _ = render_artifact(document, "markdown", metadata)
    markdown = markdown_bytes.decode("utf-8")
    assert "Nexus Artifact Delivery Packet" in markdown
    assert "Lifecycle: Candidate" in markdown
    assert "Revision: v7 · revision" in markdown
    assert "Coverage: 80% · 4/5 blocks" in markdown
    assert "Evidence: 2 bound references" in markdown
    assert "Refresh: 1 pending review" in markdown
    assert "[evidence:" not in markdown
    assert "First grounded claim[E1]" in markdown
    assert "second grounded claim[E2]" in markdown
    assert "Repeat the first source[E1]" in markdown
    assert markdown.count("- [E1]") == 1
    assert "Characters 10-28" in markdown
    assert "Evidence 019f7e45/6c41" in markdown
    assert "Evidence 019f7e45/6c99" in markdown

    html_bytes, _, _ = render_artifact(document, "html", metadata)
    html = html_bytes.decode("utf-8")
    assert 'class="delivery-cover"' in html
    assert "Nexus Artifact Delivery Packet" in html
    assert "<dt>Lifecycle</dt><dd>Candidate</dd>" in html
    assert "<dt>Coverage</dt><dd>80% · 4/5 blocks</dd>" in html
    assert "[evidence:" not in html
    assert 'class="citation" href="#source-e1"' in html
    assert 'id="source-e1"' in html
    assert "Source receipts" in html
    assert "Bounding box 121, 83, 878, 414" in html

    pdf_bytes, _, _ = render_artifact(document, "pdf", metadata)
    assert pdf_bytes.startswith(b"%PDF")


def test_artifact_html_pdf_exports_render_markdown_structure() -> None:
    evidence_id = "019f7e45-4278-7c9c-9125-9d4182f72cd8"
    document = {
        "schema": "nexus.block-document.v1",
        "title": "Structured Delivery",
        "blocks": [
            {"type": "heading", "level": 1, "text": "Structured Delivery"},
            {
                "type": "paragraph",
                "text": (
                    "Intro with **bold signal** and `inline code`"
                    f"[evidence:{evidence_id}].\n\n"
                    "---\n\n"
                    "### Inspection steps\n\n"
                    "- Capture source\n"
                    "- Render citations\n\n"
                    "1. Review export\n"
                    "2. Ship packet\n\n"
                    "| Metric | Value |\n"
                    "| --- | --- |\n"
                    "| Coverage | **100%** |\n"
                ),
                "evidence_revision_ids": [evidence_id],
            },
            {
                "type": "evidence_list",
                "items": [{"evidence_revision_id": evidence_id, "source": "architecture.md"}],
            },
        ],
    }

    html_bytes, _, _ = render_artifact(document, "html")
    html = html_bytes.decode("utf-8")

    assert "### Inspection steps" not in html
    assert "**bold signal**" not in html
    assert "<strong>bold signal</strong>" in html
    assert "<code>inline code</code>" in html
    assert "<hr>" in html
    assert 'class="delivery-map"' in html
    assert 'href="#structured-delivery"' in html
    assert 'href="#inspection-steps"' in html
    assert 'href="#source-receipts"' in html
    assert 'id="structured-delivery"' in html
    assert 'id="inspection-steps"' in html
    assert '<h3 id="inspection-steps">Inspection steps</h3>' in html
    assert "<ul><li>Capture source</li><li>Render citations</li></ul>" in html
    assert "<ol><li>Review export</li><li>Ship packet</li></ol>" in html
    assert "<table><thead><tr><th>Metric</th><th>Value</th>" in html
    assert "<td>Coverage</td><td><strong>100%</strong></td>" in html
    assert 'class="citation" href="#source-e1"' in html
    assert "@media print" in html

    pdf_bytes, _, _ = render_artifact(document, "pdf")
    assert pdf_bytes.startswith(b"%PDF")


def test_scheduled_news_reuses_connector_ingestion_boundary(tmp_path: Path) -> None:
    legacy = NexusSettings(
        _env_file=None,
        TAVILY_HOT_TOPICS_KB_ID="legacy-space",
        TAVILY_HOT_TOPICS_DEFAULT_QUERY="legacy-query",
    )
    assert legacy.scheduled_news_space_id == "legacy-space"
    assert legacy.scheduled_news_query == "legacy-query"
    settings = NexusSettings.for_test(tmp_path).model_copy(
        update={
            "scheduled_news_space_id": "space-daily",
            "scheduled_news_query": "AI research",
            "scheduled_news_topic": "news",
            "scheduled_news_time_range": "day",
            "scheduled_news_max_results": 7,
            "scheduled_news_include_full_content": True,
        }
    )

    class RecordingConnector:
        def __init__(self) -> None:
            self.call: dict[str, object] | None = None

        def sync(self, **kwargs: object) -> list[SimpleNamespace]:
            self.call = kwargs
            return [SimpleNamespace(job=SimpleNamespace(id="job-news"))]

    connector = RecordingConnector()
    result = _sync_scheduled_news(SimpleNamespace(settings=settings, connectors=connector))
    assert result == {
        "status": "scheduled",
        "space_id": "space-daily",
        "items": 1,
        "job_ids": ["job-news"],
    }
    assert connector.call == {
        "kind": "news",
        "space_id": "space-daily",
        "process_inline": False,
        "query": "AI research",
        "topic": "news",
        "time_range": "day",
        "search_depth": "advanced",
        "include_full_content": True,
        "max_results": 7,
    }

    disabled = _sync_scheduled_news(
        SimpleNamespace(
            settings=settings.model_copy(update={"scheduled_news_space_id": None}),
            connectors=connector,
        )
    )
    assert disabled["status"] == "disabled"
